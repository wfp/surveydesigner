from io import BytesIO

from accounts.const import PermissionGroups
from change_requests.const import StatusType
from change_requests.models import ChangeRequest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.files.base import ContentFile
from django.shortcuts import reverse
from django.test import TestCase
from modules.models import (
    Indicator,
    Module,
    Submodule,
    SubmoduleMapping,
    SubmoduleRequiredGroup,
)
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from organization.models import Organization
from questions.const import QuestionType
from questions.models import (
    BaseQuestion,
    Choice,
    ChoiceGroup,
    RecallPeriod,
    RootQuestion,
    SubQuestion,
    Suffix,
)
from questions.services import QuestionsExport


def test_submit_change_request_view_get(
    logged_admin_client,
    change_request_1,
):
    url = reverse("submit_change_request")
    response = logged_admin_client.get(url)
    assert response.status_code == 200


def test_submit_single_organization_change_request_view_post(
    admin, logged_admin_client
):
    description = "test description"
    admin_organization_id = 1
    admin.groups.add(Group.objects.get(name=PermissionGroups.CHANGE_REQUESTS))
    admin.save()
    with open(
        "./survey_designer/apps/questions/tests/files/questions.xlsx", "rb"
    ) as file:
        url = reverse("submit_change_request")
        response = logged_admin_client.post(
            url,
            {
                "file": file,
                "description": description,
                "organizations": [admin_organization_id],
            },
            follow=True,
        )
        assert response.status_code == 200
        created_cr = ChangeRequest.objects.last()
        assert created_cr.description == description
        assert created_cr.created_by == admin
        assert created_cr.organizations.all().count() == 1
        assert created_cr.organizations.first().id == admin_organization_id


def test_submit_multi_organization_change_request_view_post(admin, logged_admin_client):
    description = "test description"
    organization_ids = [1, 2]
    admin.groups.add(Group.objects.get(name=PermissionGroups.CHANGE_REQUESTS))
    admin.save()
    with open(
        "./survey_designer/apps/questions/tests/files/questions.xlsx", "rb"
    ) as file:
        url = reverse("submit_change_request")
        response = logged_admin_client.post(
            url,
            {
                "file": file,
                "description": description,
                "organizations": organization_ids,
            },
            follow=True,
        )
        assert response.status_code == 200
        created_cr = ChangeRequest.objects.last()
        assert created_cr.description == description
        assert created_cr.created_by == admin
        assert created_cr.organizations.all().count() == len(organization_ids)
        assert (
            list(created_cr.organizations.all().values_list("id", flat=True))
            == organization_ids
        )


def test_approve_change_request_view_get(logged_admin_client, change_request_1):
    url = reverse("approve_change_request", args=[change_request_1.id])
    response = logged_admin_client.get(url)
    assert response.status_code == 200


def test_approve_change_request_view_post(logged_admin_client, change_request_1):
    cr_response = "test response"
    url = reverse("approve_change_request", args=[change_request_1.id])
    response = logged_admin_client.post(url, {"response": cr_response}, follow=True)
    assert response.status_code == 200
    change_request_1.refresh_from_db()
    assert change_request_1.status == StatusType.APPROVED
    assert change_request_1.response == cr_response


class TestApproveChangeRequestPreview(TestCase):
    def setUp(self):
        self.admin = get_user_model().objects.create_user(
            email="preview-admin@example.com",
            password="admin_user",
            is_superuser=True,
            is_staff=True,
        )
        self.client.force_login(self.admin)

        self.organization = Organization.objects.create(name="Preview Organization")
        module = Module.objects.create(name="PreviewModule", label="Preview Module")
        module.organizations.add(self.organization)
        self.submodule = Submodule.objects.create(
            name="PreviewSubmodule",
            label="Preview Submodule",
            module=module,
            mapping=SubmoduleMapping.objects.create(),
        )

        self.choice_group = ChoiceGroup.objects.create(name="PreviewChoices")
        Choice.objects.create(
            choice_group=self.choice_group,
            name="1",
            label="Yes",
            order=1,
        )
        Choice.objects.create(
            choice_group=self.choice_group,
            name="0",
            label="No",
            order=2,
        )

        self.root_question = RootQuestion.objects.create(
            name="PreviewRootQuestion",
            label="Preview Root Question",
            type=QuestionType.INTEGER,
            description="Original root description",
        )
        self.root_question.submodule.add(self.submodule)

        self.choice_root_question = RootQuestion.objects.create(
            name="PreviewChoiceQuestion",
            label="Preview Choice Question",
            type=QuestionType.SELECT_MULTIPLE,
            choices=self.choice_group,
            description="Original choice description",
        )
        self.choice_root_question.submodule.add(self.submodule)

        self.suffix = Suffix.objects.create(
            name="_PreviewSuffix",
            description="Preview suffix",
            type=QuestionType.SELECT_ONE,
            choices=self.choice_group,
        )
        self.recall_period = RecallPeriod.objects.create(
            name="_PreviewRecall",
            description="Preview recall",
        )

        self.suffix_sub_question = SubQuestion.objects.create(
            root_question=self.root_question,
            name="PreviewSuffixSubQuestion",
            suffix=self.suffix,
            label="Preview suffix sub question",
            description="Original suffix description",
        )
        self.recall_sub_question = SubQuestion.objects.create(
            root_question=self.root_question,
            name="PreviewRecallSubQuestion",
            recall_period=self.recall_period,
            label="Preview recall sub question",
            description="Original recall description",
        )

        indicator = Indicator.objects.create(
            name="PreviewIndicator",
            label="Preview Indicator",
        )
        indicator.questions.add(
            self.choice_root_question.base_question,
            self.suffix_sub_question.base_question,
            self.recall_sub_question.base_question,
        )

        SubmoduleRequiredGroup.objects.create(
            submodule=self.submodule,
            required_suffix=self.suffix,
        )
        SubmoduleRequiredGroup.objects.create(
            submodule=self.submodule,
            required_recall_period=self.recall_period,
        )

    def test_existing_exported_questions_only_show_question_updates(self):
        target_base_questions = BaseQuestion.objects.filter(
            id__in=[
                self.choice_root_question.base_question.id,
                self.suffix_sub_question.base_question.id,
                self.recall_sub_question.base_question.id,
            ]
        )
        target_names = {
            self.choice_root_question.name,
            self.suffix_sub_question.name,
            self.recall_sub_question.name,
        }

        workbook = load_workbook(
            BytesIO(QuestionsExport().generate_from_questions(target_base_questions))
        )
        survey_sheet = workbook["survey"]
        headers = [cell.value for cell in survey_sheet[1]]
        name_column = headers.index("name") + 1
        description_column = headers.index("description") + 1

        for row_number in range(2, survey_sheet.max_row + 1):
            question_name = survey_sheet.cell(row=row_number, column=name_column).value
            if question_name not in target_names:
                continue

            current_description = (
                survey_sheet.cell(row=row_number, column=description_column).value or ""
            )
            survey_sheet.cell(
                row=row_number,
                column=description_column,
                value=f"{current_description} 1".strip(),
            )

        change_request = ChangeRequest.objects.create(
            created_by=self.admin,
            description="Preview existing questions",
        )
        change_request.file.save(
            "preview_existing_questions.xlsx",
            ContentFile(save_virtual_workbook(workbook)),
        )
        change_request.organizations.add(self.organization)

        response = self.client.get(
            reverse("approve_change_request", args=[change_request.id])
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["other_errors"], {})

        updates = response.context["updates"]
        self.assertEqual(set(updates.keys()), {"Questions Spreadsheet"})
        self.assertEqual(updates["Questions Spreadsheet"]["To Create"], {})
        self.assertEqual(
            set(updates["Questions Spreadsheet"]["To Update"]["Question"]),
            target_names,
        )

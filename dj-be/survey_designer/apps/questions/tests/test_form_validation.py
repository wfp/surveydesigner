import io

import pytest
from django.core.files.base import ContentFile
from openpyxl import Workbook
from questions.models import ChoiceGroup
from questions.services import QuestionsExport, XLSForm
from questions.services.form_validation import (
    ArtifactInputError,
    GeneratedSurveyArtifact,
    ValidationIssue,
    build_generated_artifact,
    compute_artifact_hash,
    materialize_external_files,
    validate_codebook_integrity,
    validate_generated_artifact,
    validate_xml_compatibility,
)
from questions.services.xml_conversion import XMLConversion


def _codebook_workbook(
    survey_rows,
    choice_rows=(),
    *,
    export_columns=False,
    label_column="label",
    suffix_rows=(),
    recall_period_rows=(),
    survey_columns=None,
    choice_columns=None,
):
    workbook = Workbook()
    survey = workbook.active
    survey.title = "survey"
    default_survey_columns = (
        [
            "type",
            "name",
            "choice_list",
            "suffix1",
            "suffix2",
            "recall_period",
        ]
        if export_columns
        else ["type", "name"]
    )
    survey.append(survey_columns or default_survey_columns)
    for row in survey_rows:
        survey.append(row)

    choices = workbook.create_sheet("choices")
    default_choice_columns = (
        ["choice_list", "name", label_column]
        if export_columns
        else ["list_name", "name", label_column]
    )
    choices.append(choice_columns or default_choice_columns)
    for row in choice_rows:
        choices.append(row)

    if export_columns or suffix_rows:
        suffixes = workbook.create_sheet("suffixes")
        suffixes.append(["name", "description", "type", "choicelist", "suffix"])
        for row in suffix_rows:
            suffixes.append(row)

    if export_columns or recall_period_rows:
        recall_periods = workbook.create_sheet("recall_periods")
        recall_periods.append(["name", "description"])
        for row in recall_period_rows:
            recall_periods.append(row)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _choice_filter_workbook(
    expression,
    *,
    source_question="region",
    source_values=("north",),
    emitted_filter_column="choice_filter_name",
    filter_values=("north",),
):
    choice_rows = [("regions", value, "", f"Region {value}") for value in source_values]
    choice_rows.extend(
        ("cities", f"city_{index}", value, f"City {index}")
        for index, value in enumerate(filter_values, start=1)
    )
    return _codebook_workbook(
        [
            ("select_one regions", source_question, ""),
            ("select_one cities", "city", expression),
        ],
        choice_rows,
        survey_columns=["type", "name", "choice_filter"],
        choice_columns=[
            "list_name",
            "name",
            emitted_filter_column,
            "label",
        ],
    )


def _configure_generated_choice_filter(
    source_question,
    filtered_question,
    source_choices,
    filtered_choices,
):
    source_question.type = "select_one"
    source_question.choices = source_choices
    source_question.save(update_fields=["type", "choices"])
    filtered_question.type = "select_one"
    filtered_question.choices = filtered_choices
    filtered_question.choice_filter = f"choice_filter_name=${{{source_question.name}}}"
    filtered_question.save(update_fields=["type", "choices", "choice_filter"])
    filtered_choices.choice_filter_list = source_choices
    filtered_choices.save(update_fields=["choice_filter_list"])

    source_values = list(source_choices.choices.order_by("order", "id"))
    for choice, source_value in zip(
        filtered_choices.choices.order_by("order", "id"), source_values
    ):
        choice.choice_filter_name = source_value
        choice.save(update_fields=["choice_filter_name"])


class ConversionMustNotRun:
    def __init__(self, xlsx_file):
        raise AssertionError("pyxform must not run for an invalid codebook")


def test_artifact_hash_is_order_independent_for_external_files():
    first = compute_artifact_hash(b"xlsx", {"b.csv": b"2", "a.csv": b"1"})
    second = compute_artifact_hash(b"xlsx", {"a.csv": b"1", "b.csv": b"2"})

    assert first == second
    assert first.startswith("sha256:")
    assert first != compute_artifact_hash(b"different", {"a.csv": b"1", "b.csv": b"2"})


def test_build_generated_artifact_generates_and_reads_each_external_file_once():
    class ReadOnceFile:
        def __init__(self, content):
            self.content = content
            self.open_mode = None
            self.read_count = 0
            self.close_count = 0

        def open(self, mode):
            self.open_mode = mode
            return self

        def read(self):
            self.read_count += 1
            return self.content

        def close(self):
            self.close_count += 1

    external = ReadOnceFile(b"name\nvalue\n")

    class Form:
        id_name = "generated-form"

        def __init__(self):
            self.generate_count = 0
            self.external_files = {"choices.csv": external}

        def generate(self):
            self.generate_count += 1
            return b"xlsx-bytes"

    form = Form()
    artifact = build_generated_artifact(form)

    assert artifact.xlsx_bytes == b"xlsx-bytes"
    assert artifact.external_files == {"choices.csv": b"name\nvalue\n"}
    assert form.generate_count == 1
    assert external.open_mode == "rb"
    assert external.read_count == 1
    assert external.close_count == 1


def test_compatibility_requires_referenced_csv_in_exact_artifact():
    xml = (
        '<h:html xmlns:h="http://www.w3.org/1999/xhtml" '
        'xmlns:xf="http://www.w3.org/2002/xforms">'
        "<h:head><xf:model><xf:instance><data/></xf:instance>"
        '<xf:instance id="choices" src="jr://file-csv/choices.csv"/>'
        "</xf:model></h:head><h:body/></h:html>"
    )

    issues = validate_xml_compatibility(xml)
    assert [issue.code for issue in issues] == ["EXTERNAL_FILE_MISSING"]
    assert validate_xml_compatibility(xml, {"choices.csv": b"data"}) == []


def test_validation_result_normalizes_compatibility_issues():
    class Conversion:
        def __init__(self, xlsx_file):
            self.xlsx_file = xlsx_file
            self.errors = []
            self.warnings = ["non-blocking warning"]

        def run(self):
            return (
                '<h:html xmlns:h="http://www.w3.org/1999/xhtml" '
                'xmlns:xf="http://www.w3.org/2002/xforms">'
                "<h:head><xf:model><xf:instance><data/></xf:instance>"
                "</xf:model></h:head><h:body/></h:html>"
            )

    artifact = GeneratedSurveyArtifact(_codebook_workbook([("text", "notes")]))
    result = validate_generated_artifact(artifact, converter_cls=Conversion)

    assert result.valid is True
    assert result.errors == ()
    assert result.warnings[0].as_dict() == {
        "code": "PYXFORM_WARNING",
        "layer": "pyxform",
        "severity": "warning",
        "message": "non-blocking warning",
    }
    assert result.as_dict()["validator"] == {
        "pyxform": "4.5.0",
        "compatibility": "1.0",
    }


@pytest.mark.parametrize("question_type", ["select_one", "select_multiple"])
def test_codebook_validation_accepts_internal_select_with_emitted_choices(
    question_type,
):
    xlsx = _codebook_workbook(
        [(f"{question_type} foods", "preferred_food")],
        [("foods", "rice", "Rice")],
    )

    assert validate_codebook_integrity(xlsx) == []


def test_codebook_validation_reports_internal_select_without_choice_list():
    xlsx = _codebook_workbook([("select_one", "preferred_food")])

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_CHOICE_LIST_MISSING"]
    assert issues[0].as_dict() == {
        "code": "CODEBOOK_CHOICE_LIST_MISSING",
        "layer": "composition",
        "severity": "error",
        "message": "Question 'preferred_food' uses select_one but does not specify a choice list.",
        "owner": {"model": "question", "name": "preferred_food"},
        "field": "choices",
        "sheet": "survey",
        "column": "type",
        "row": 2,
    }


def test_codebook_validation_reports_choice_list_without_emitted_rows():
    xlsx = _codebook_workbook([("select_multiple foods", "preferred_foods")])

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_CHOICE_LIST_NOT_EMITTED"]
    assert "choice list 'foods'" in issues[0].message


def test_codebook_validation_uses_export_choice_list_column():
    xlsx = _codebook_workbook(
        [("select_one", "preferred_food", "foods")],
        [("foods", "rice", "Rice")],
        export_columns=True,
    )

    assert validate_codebook_integrity(xlsx) == []


@pytest.mark.parametrize("missing_value", [None, "", "   "])
def test_codebook_validation_reports_missing_choice_value(missing_value):
    xlsx = _codebook_workbook(
        [("select_one foods", "preferred_food")],
        [("foods", missing_value, "Rice")],
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_CHOICE_VALUE_MISSING"]
    assert issues[0].as_dict() == {
        "code": "CODEBOOK_CHOICE_VALUE_MISSING",
        "layer": "composition",
        "severity": "error",
        "message": "Choice list 'foods' has an emitted choice with no value.",
        "owner": {"model": "choice_list", "name": "foods"},
        "field": "name",
        "sheet": "choices",
        "column": "name",
        "row": 2,
    }


@pytest.mark.parametrize("missing_label", [None, "", "   "])
@pytest.mark.parametrize("label_column", ["label", "label::English (en)"])
def test_codebook_validation_reports_missing_english_label(
    missing_label,
    label_column,
):
    xlsx = _codebook_workbook(
        [("select_one foods", "preferred_food")],
        [("foods", "rice", missing_label)],
        label_column=label_column,
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_CHOICE_LABEL_MISSING"]
    assert issues[0].as_dict() == {
        "code": "CODEBOOK_CHOICE_LABEL_MISSING",
        "layer": "composition",
        "severity": "error",
        "message": "Choice 'rice' in choice list 'foods' has no English label.",
        "owner": {"model": "choice", "name": "rice", "choice_list": "foods"},
        "field": "label",
        "sheet": "choices",
        "column": label_column,
        "row": 2,
    }


def test_codebook_validation_accepts_numeric_choice_value():
    xlsx = _codebook_workbook(
        [("select_one foods", "preferred_food")],
        [("foods", 0, "None")],
    )

    assert validate_codebook_integrity(xlsx) == []


def test_codebook_validation_reports_duplicate_value_in_same_list():
    xlsx = _codebook_workbook(
        [("select_one foods", "preferred_food")],
        [
            ("foods", "rice", "Rice"),
            ("foods", "rice", "Rice duplicate"),
        ],
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_CHOICE_VALUE_DUPLICATE"]
    assert issues[0].as_dict() == {
        "code": "CODEBOOK_CHOICE_VALUE_DUPLICATE",
        "layer": "composition",
        "severity": "error",
        "message": "Choice value 'rice' is duplicated in choice list 'foods'; it was first emitted at row 2.",
        "owner": {"model": "choice", "name": "rice", "choice_list": "foods"},
        "field": "name",
        "sheet": "choices",
        "column": "name",
        "row": 3,
    }


def test_codebook_validation_reports_every_duplicate_after_first_row():
    xlsx = _codebook_workbook(
        [("select_one foods", "preferred_food")],
        [
            ("foods", "rice", "Rice"),
            ("foods", "rice", "Rice duplicate"),
            ("foods", "rice", "Rice duplicate again"),
        ],
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.row for issue in issues] == [3, 4]
    assert all("first emitted at row 2" in issue.message for issue in issues)


def test_codebook_validation_allows_same_value_in_different_lists():
    xlsx = _codebook_workbook(
        [
            ("select_one foods", "preferred_food"),
            ("select_one drinks", "preferred_drink"),
        ],
        [
            ("foods", "other", "Other food"),
            ("drinks", "other", "Other drink"),
        ],
    )

    assert validate_codebook_integrity(xlsx) == []


def test_codebook_validation_compares_choice_values_case_sensitively():
    xlsx = _codebook_workbook(
        [("select_one foods", "preferred_food")],
        [
            ("foods", "Other", "Uppercase"),
            ("foods", "other", "Lowercase"),
        ],
    )

    assert validate_codebook_integrity(xlsx) == []


@pytest.mark.parametrize(
    "first_value, duplicate_value",
    [(" rice ", "rice"), (1, "1")],
)
def test_codebook_validation_uses_emitted_value_normalization_for_duplicates(
    first_value,
    duplicate_value,
):
    xlsx = _codebook_workbook(
        [("select_one foods", "preferred_food")],
        [
            ("foods", first_value, "First"),
            ("foods", duplicate_value, "Duplicate"),
        ],
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_CHOICE_VALUE_DUPLICATE"]


def test_duplicate_choice_value_blocks_pyxform_conversion():
    artifact = GeneratedSurveyArtifact(
        _codebook_workbook(
            [("select_one foods", "preferred_food")],
            [
                ("foods", "rice", "Rice"),
                ("foods", "rice", "Rice duplicate"),
            ],
        )
    )

    result = validate_generated_artifact(artifact, converter_cls=ConversionMustNotRun)

    assert result.valid is False
    assert [issue.code for issue in result.errors] == [
        "CODEBOOK_CHOICE_VALUE_DUPLICATE"
    ]


def test_codebook_export_layout_reports_duplicate_choice_value():
    xlsx = _codebook_workbook(
        [("select_one", "preferred_food", "foods")],
        [
            ("foods", "rice", "Rice"),
            ("foods", "rice", "Rice duplicate"),
        ],
        export_columns=True,
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_CHOICE_VALUE_DUPLICATE"]


@pytest.mark.parametrize(
    "list_name",
    ["foods", "_foods", "food-list", "food.list", "food_1", "éfoods"],
)
def test_codebook_validation_accepts_compatible_choice_list_names(list_name):
    xlsx = _codebook_workbook(
        [(f"select_one {list_name}", "preferred_food")],
        [(list_name, "rice", "Rice")],
    )

    assert validate_codebook_integrity(xlsx) == []


@pytest.mark.parametrize(
    "list_name",
    ["1foods", "food list", "food$list", "food/list", "food:items"],
)
def test_codebook_validation_reports_invalid_choice_list_name(list_name):
    xlsx = _codebook_workbook(
        [(f"select_one {list_name}", "preferred_food")],
        [(list_name, "rice", "Rice")],
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_CHOICE_LIST_NAME_INVALID"]
    assert issues[0].as_dict() == {
        "code": "CODEBOOK_CHOICE_LIST_NAME_INVALID",
        "layer": "composition",
        "severity": "error",
        "message": f"Choice list name '{list_name}' is invalid. Names must begin with a letter or underscore and contain only letters, digits, underscores, hyphens, or periods.",
        "owner": {"model": "choice_list", "name": list_name},
        "field": "list_name",
        "sheet": "choices",
        "column": "list_name",
        "row": 2,
    }


def test_invalid_unemitted_choice_list_name_is_reported_at_question():
    xlsx = _codebook_workbook([("select_one 1foods", "preferred_food")])

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == [
        "CODEBOOK_CHOICE_LIST_NAME_INVALID",
        "CODEBOOK_CHOICE_LIST_NOT_EMITTED",
    ]
    assert issues[0].sheet == "survey"
    assert issues[0].row == 2
    assert issues[0].owner == {"model": "question", "name": "preferred_food"}


def test_codebook_validation_reports_choice_row_without_list_name():
    xlsx = _codebook_workbook(
        [("select_one foods", "preferred_food")],
        [("", "rice", "Rice")],
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == [
        "CODEBOOK_CHOICE_LIST_NAME_MISSING",
        "CODEBOOK_CHOICE_LIST_NOT_EMITTED",
    ]
    assert issues[0].as_dict() == {
        "code": "CODEBOOK_CHOICE_LIST_NAME_MISSING",
        "layer": "composition",
        "severity": "error",
        "message": "An emitted choice row does not specify a choice list name.",
        "owner": {"model": "choice", "name": "rice"},
        "field": "list_name",
        "sheet": "choices",
        "column": "list_name",
        "row": 2,
    }


@pytest.mark.parametrize("choice_value", ["with space", "with\tspace"])
def test_codebook_validation_rejects_whitespace_in_select_multiple_value(
    choice_value,
):
    xlsx = _codebook_workbook(
        [("select_multiple foods", "preferred_foods")],
        [("foods", choice_value, "Value")],
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_CHOICE_VALUE_INVALID"]
    assert issues[0].as_dict() == {
        "code": "CODEBOOK_CHOICE_VALUE_INVALID",
        "layer": "composition",
        "severity": "error",
        "message": f"Choice value '{choice_value}' in choice list 'foods' contains whitespace, which is not supported by select_multiple questions.",
        "owner": {
            "model": "choice",
            "name": choice_value,
            "choice_list": "foods",
        },
        "field": "name",
        "sheet": "choices",
        "column": "name",
        "row": 2,
    }


def test_codebook_validation_allows_whitespace_in_select_one_value():
    xlsx = _codebook_workbook(
        [("select_one foods", "preferred_food")],
        [("foods", "white rice", "White rice")],
    )

    assert validate_codebook_integrity(xlsx) == []


def test_codebook_validation_uses_strictest_type_for_shared_choice_list():
    xlsx = _codebook_workbook(
        [
            ("select_one foods", "preferred_food"),
            ("select_multiple foods", "preferred_foods"),
        ],
        [("foods", "white rice", "White rice")],
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_CHOICE_VALUE_INVALID"]


@pytest.mark.parametrize(
    "choice_value",
    ["1", "1.5", "-1", "a-b", "a.b", "a/b", "a:b", "é"],
)
def test_codebook_validation_accepts_compatible_select_multiple_values(choice_value):
    xlsx = _codebook_workbook(
        [("select_multiple foods", "preferred_foods")],
        [("foods", choice_value, "Value")],
    )

    assert validate_codebook_integrity(xlsx) == []


def test_choice_value_syntax_validation_supports_codebook_export_layout():
    xlsx = _codebook_workbook(
        [("select_multiple", "preferred_foods", "foods")],
        [("foods", "white rice", "White rice")],
        export_columns=True,
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_CHOICE_VALUE_INVALID"]


def test_invalid_choice_syntax_blocks_pyxform_conversion():
    artifact = GeneratedSurveyArtifact(
        _codebook_workbook(
            [("select_multiple foods", "preferred_foods")],
            [("foods", "white rice", "White rice")],
        )
    )

    result = validate_generated_artifact(artifact, converter_cls=ConversionMustNotRun)

    assert result.valid is False
    assert [issue.code for issue in result.errors] == ["CODEBOOK_CHOICE_VALUE_INVALID"]


def test_codebook_validation_reports_missing_generated_name():
    xlsx = _codebook_workbook([("text", "")])

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_GENERATED_NAME_MISSING"]
    assert issues[0].as_dict() == {
        "code": "CODEBOOK_GENERATED_NAME_MISSING",
        "layer": "composition",
        "severity": "error",
        "message": "Survey row 2 of type 'text' requires a generated name.",
        "owner": {"model": "question", "type": "text"},
        "field": "name",
        "sheet": "survey",
        "column": "name",
        "row": 2,
    }


def test_codebook_validation_allows_unnamed_closing_rows():
    xlsx = _codebook_workbook(
        [
            ("begin_group", "household"),
            ("end_group", ""),
            ("begin_repeat", "members"),
            ("end repeat", ""),
        ]
    )

    assert validate_codebook_integrity(xlsx) == []


@pytest.mark.parametrize(
    "name",
    ["1question", "question name", "question$name", "question/name", "question:name"],
)
def test_codebook_validation_reports_invalid_generated_name(name):
    xlsx = _codebook_workbook([("text", name)])

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_GENERATED_NAME_INVALID"]
    assert issues[0].as_dict() == {
        "code": "CODEBOOK_GENERATED_NAME_INVALID",
        "layer": "composition",
        "severity": "error",
        "message": f"Generated question name '{name}' is invalid. Names must begin with a letter or underscore and contain only letters, digits, underscores, hyphens, or periods.",
        "owner": {"model": "question", "name": name, "type": "text"},
        "field": "name",
        "sheet": "survey",
        "column": "name",
        "row": 2,
    }


@pytest.mark.parametrize(
    "name",
    ["question", "_question", "question-1", "question.1", "question_1", "équestion"],
)
def test_codebook_validation_accepts_compatible_generated_name(name):
    xlsx = _codebook_workbook([("text", name)])

    assert validate_codebook_integrity(xlsx) == []


def test_codebook_validation_rejects_reserved_generated_name():
    xlsx = _codebook_workbook([("text", "meta")])

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_GENERATED_NAME_INVALID"]
    assert issues[0].message == (
        "Generated question name 'meta' is reserved and cannot be emitted on the "
        "survey sheet."
    )


def test_codebook_validation_reports_duplicate_generated_name_across_row_types():
    xlsx = _codebook_workbook(
        [
            ("begin_group", "household"),
            ("text", "household"),
            ("end_group", ""),
        ]
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_GENERATED_NAME_DUPLICATE"]
    assert issues[0].as_dict() == {
        "code": "CODEBOOK_GENERATED_NAME_DUPLICATE",
        "layer": "composition",
        "severity": "error",
        "message": "Generated question name 'household' duplicates a group first emitted at survey row 2.",
        "owner": {
            "model": "question",
            "name": "household",
            "type": "text",
            "first_model": "group",
            "first_row": 2,
        },
        "field": "name",
        "sheet": "survey",
        "column": "name",
        "row": 3,
    }


def test_codebook_validation_reports_every_generated_name_duplicate_after_first():
    xlsx = _codebook_workbook(
        [("text", "shared"), ("integer", "shared"), ("decimal", "shared")]
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.row for issue in issues] == [3, 4]
    assert all(issue.code == "CODEBOOK_GENERATED_NAME_DUPLICATE" for issue in issues)
    assert all("first emitted at survey row 2" in issue.message for issue in issues)


def test_codebook_validation_compares_generated_names_case_sensitively():
    xlsx = _codebook_workbook([("text", "Household"), ("text", "household")])

    assert validate_codebook_integrity(xlsx) == []


def test_questions_export_layout_does_not_apply_final_name_collision_rule():
    xlsx = _codebook_workbook(
        [("text", "shared", ""), ("text", "shared", "")],
        export_columns=True,
    )

    assert validate_codebook_integrity(xlsx) == []


def test_invalid_generated_name_blocks_pyxform_conversion():
    artifact = GeneratedSurveyArtifact(_codebook_workbook([("text", "1invalid")]))

    result = validate_generated_artifact(artifact, converter_cls=ConversionMustNotRun)

    assert result.valid is False
    assert [issue.code for issue in result.errors] == [
        "CODEBOOK_GENERATED_NAME_INVALID"
    ]


def test_codebook_export_accepts_valid_suffix_and_recall_period_composition():
    xlsx = _codebook_workbook(
        [("select_one", "food_adult_male_7d", "yes_no", "_adult", "_male", "_7d")],
        [("yes_no", "yes", "Yes")],
        export_columns=True,
        suffix_rows=[
            ("_adult", "Adult", "text", "", "_male"),
            ("_male", "Male", "select_one", "yes_no", ""),
        ],
        recall_period_rows=[("_7d", "Last seven days")],
    )

    assert validate_codebook_integrity(xlsx) == []


def test_real_codebook_export_with_suffix_and_recall_period_passes(
    sub_question_1,
    sub_question_2,
):
    base_question_model = type(sub_question_1.base_question)
    export = QuestionsExport(languages=[("en", "English")])
    xlsx = export.generate_from_questions(
        base_question_model.objects.filter(
            id__in=[
                sub_question_1.base_question.id,
                sub_question_2.base_question.id,
            ]
        )
    )

    assert validate_codebook_integrity(xlsx) == []


@pytest.mark.parametrize(
    "survey_row, expected_code, expected_field",
    [
        (
            ("text", "food_missing", "", "_missing", "", ""),
            "CODEBOOK_SUFFIX_NOT_EMITTED",
            "suffix1",
        ),
        (
            ("text", "food_7d", "", "", "", "_7d"),
            "CODEBOOK_RECALL_PERIOD_NOT_EMITTED",
            "recall_period",
        ),
    ],
)
def test_codebook_export_reports_non_emitted_generated_reference(
    survey_row,
    expected_code,
    expected_field,
):
    xlsx = _codebook_workbook([survey_row], export_columns=True)

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == [expected_code]
    assert issues[0].field == expected_field
    assert issues[0].owner == {"model": "question", "name": survey_row[1]}


def test_codebook_export_reports_suffix_2_without_suffix_1():
    xlsx = _codebook_workbook(
        [("text", "food_male", "", "", "_male", "")],
        export_columns=True,
        suffix_rows=[("_male", "Male", "text", "", "")],
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_NESTED_SUFFIX_INVALID"]
    assert "without suffix 1" in issues[0].message


def test_codebook_export_reports_incompatible_nested_suffix():
    xlsx = _codebook_workbook(
        [("text", "food_adult_male", "", "_adult", "_male", "")],
        export_columns=True,
        suffix_rows=[
            ("_adult", "Adult", "text", "", "_female"),
            ("_male", "Male", "text", "", ""),
        ],
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_NESTED_SUFFIX_INVALID"]
    assert "not nested under suffix 1 '_adult'" in issues[0].message


@pytest.mark.parametrize(
    "suffix_row, expected_code",
    [
        (("_adult", "Adult", "same", "", ""), "CODEBOOK_SUFFIX_TYPE_UNSUPPORTED"),
        (
            ("_adult", "Adult", "select_one", "", ""),
            "CODEBOOK_SUFFIX_CHOICE_LIST_MISSING",
        ),
        (
            ("_adult", "Adult", "select_one", "missing", ""),
            "CODEBOOK_SUFFIX_CHOICE_LIST_NOT_EMITTED",
        ),
        (
            ("_adult", "Adult", "text", "yes_no", ""),
            "CODEBOOK_SUFFIX_CHOICE_LIST_INCOMPATIBLE",
        ),
    ],
)
def test_codebook_export_reports_invalid_suffix_definition(
    suffix_row,
    expected_code,
):
    xlsx = _codebook_workbook(
        [],
        export_columns=True,
        suffix_rows=[suffix_row],
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == [expected_code]
    assert issues[0].owner["model"] == "suffix"


def test_codebook_export_reports_effective_suffix_type_mismatch():
    xlsx = _codebook_workbook(
        [("integer", "food_adult", "", "_adult", "", "")],
        export_columns=True,
        suffix_rows=[("_adult", "Adult", "text", "", "")],
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_SUFFIX_TYPE_INCOMPATIBLE"]


def test_codebook_export_reports_effective_suffix_choice_list_mismatch():
    xlsx = _codebook_workbook(
        [("select_one", "food_adult", "list_b", "_adult", "", "")],
        [("list_a", "yes", "Yes"), ("list_b", "yes", "Yes")],
        export_columns=True,
        suffix_rows=[("_adult", "Adult", "select_one", "list_a", "")],
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == [
        "CODEBOOK_SUFFIX_CHOICE_LIST_INCOMPATIBLE"
    ]


@pytest.mark.parametrize(
    "suffix_rows, expected_code",
    [
        (
            [("", "Description", "text", "", "")],
            "CODEBOOK_SUFFIX_NAME_MISSING",
        ),
        (
            [("1invalid", "Description", "text", "", "")],
            "CODEBOOK_SUFFIX_NAME_INVALID",
        ),
        (
            [
                ("_adult", "First", "text", "", ""),
                ("_adult", "Second", "text", "", ""),
            ],
            "CODEBOOK_SUFFIX_NAME_DUPLICATE",
        ),
    ],
)
def test_codebook_export_reports_invalid_suffix_name(suffix_rows, expected_code):
    xlsx = _codebook_workbook(
        [],
        export_columns=True,
        suffix_rows=suffix_rows,
    )

    assert [issue.code for issue in validate_codebook_integrity(xlsx)] == [
        expected_code
    ]


@pytest.mark.parametrize(
    "recall_rows, expected_code",
    [
        ([("", "Last seven days")], "CODEBOOK_RECALL_PERIOD_NAME_MISSING"),
        ([("7 days", "Last seven days")], "CODEBOOK_RECALL_PERIOD_NAME_INVALID"),
        (
            [("_7d", "First"), ("_7d", "Second")],
            "CODEBOOK_RECALL_PERIOD_NAME_DUPLICATE",
        ),
    ],
)
def test_codebook_export_reports_invalid_recall_period_name(
    recall_rows,
    expected_code,
):
    xlsx = _codebook_workbook(
        [],
        export_columns=True,
        recall_period_rows=recall_rows,
    )

    assert [issue.code for issue in validate_codebook_integrity(xlsx)] == [
        expected_code
    ]


def test_final_artifact_reports_unsupported_generated_type():
    xlsx = _codebook_workbook([("same", "food_adult")])

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_GENERATED_TYPE_UNSUPPORTED"]
    assert issues[0].owner == {
        "model": "question",
        "name": "food_adult",
        "type": "same",
    }


def test_final_artifact_reports_missing_generated_type():
    xlsx = _codebook_workbook([("", "food_adult")])

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_GENERATED_TYPE_UNSUPPORTED"]
    assert "has no type" in issues[0].message


def test_final_artifact_uses_source_context_for_nested_suffix_validation():
    xlsx = _codebook_workbook([("text", "food_adult_male")])
    source_map = {
        ("survey", 2): {
            "model": "SubQuestion",
            "id": 42,
            "name": "food_adult_male",
            "suffix": {
                "model": "Suffix",
                "id": 10,
                "name": "_adult",
                "nested_suffixes": (),
            },
            "suffix_2": {"model": "Suffix", "id": 11, "name": "_male"},
            "recall_period": None,
            "effective_type": "text",
            "effective_choice_list": "",
        }
    }

    issues = validate_codebook_integrity(xlsx, source_map)

    assert [issue.code for issue in issues] == ["CODEBOOK_NESTED_SUFFIX_INVALID"]
    assert issues[0].owner == {
        "model": "SubQuestion",
        "id": 42,
        "name": "food_adult_male",
    }


def test_final_artifact_reports_invalid_suffix_name_with_suffix_owner():
    xlsx = _codebook_workbook([("text", "food_adult")])
    source_map = {
        ("survey", 2): {
            "model": "SubQuestion",
            "id": 42,
            "name": "food_adult",
            "suffix": {
                "model": "Suffix",
                "id": 10,
                "name": "bad suffix",
                "nested_suffixes": (),
            },
            "suffix_2": None,
            "recall_period": None,
            "effective_type": "text",
            "effective_choice_list": "",
        }
    }

    issues = validate_codebook_integrity(xlsx, source_map)

    assert [issue.code for issue in issues] == ["CODEBOOK_SUFFIX_NAME_INVALID"]
    assert issues[0].owner == {
        "model": "Suffix",
        "id": 10,
        "name": "bad suffix",
    }


def test_codebook_validation_does_not_decide_non_english_label_fallback():
    xlsx = _codebook_workbook(
        [("select_one foods", "preferred_food")],
        [("foods", "rice", "")],
        label_column="label::French (fr)",
    )

    assert validate_codebook_integrity(xlsx) == []


def test_codebook_validation_ignores_external_selects():
    xlsx = _codebook_workbook([("select_one_from_file foods.csv", "preferred_food")])

    assert validate_codebook_integrity(xlsx) == []


@pytest.mark.parametrize(
    "expression",
    [
        "choice_filter_name=${region}",
        "${region}=choice_filter_name",
        "selected(${region}, choice_filter_name)",
    ],
)
def test_codebook_validation_accepts_emitted_choice_filter_schema(expression):
    xlsx = _choice_filter_workbook(expression)

    assert validate_codebook_integrity(xlsx) == []


def test_codebook_validation_reports_unknown_choice_filter_column():
    xlsx = _choice_filter_workbook("unknown_column=${region}")

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == [
        "CODEBOOK_CHOICE_FILTER_COLUMN_NOT_EMITTED"
    ]
    assert issues[0].owner == {
        "model": "question",
        "name": "city",
        "type": "select_one cities",
    }


def test_codebook_validation_reports_choice_filter_column_case_mismatch():
    xlsx = _choice_filter_workbook(
        "choice_filter_name=${region}",
        emitted_filter_column="Choice_Filter_Name",
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == [
        "CODEBOOK_CHOICE_FILTER_COLUMN_CASE_MISMATCH"
    ]
    assert "available column: 'Choice_Filter_Name'" in issues[0].message


def test_codebook_validation_reports_empty_choice_filter_column():
    xlsx = _choice_filter_workbook("choice_filter_name=${region}", filter_values=("",))

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == [
        "CODEBOOK_CHOICE_FILTER_COLUMN_VALUES_MISSING"
    ]


def test_codebook_validation_reports_unavailable_choice_filter_question():
    xlsx = _choice_filter_workbook("choice_filter_name=${not_selected}")

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == [
        "CODEBOOK_CHOICE_FILTER_QUESTION_NOT_EMITTED"
    ]


def test_codebook_validation_reports_choice_filter_question_case_mismatch():
    xlsx = _choice_filter_workbook(
        "choice_filter_name=${region}", source_question="Region"
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == [
        "CODEBOOK_CHOICE_FILTER_QUESTION_CASE_MISMATCH"
    ]
    assert "available question: 'Region'" in issues[0].message


def test_codebook_validation_reports_unavailable_choice_filter_value():
    xlsx = _choice_filter_workbook(
        "choice_filter_name=${region}", filter_values=("south",)
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == [
        "CODEBOOK_CHOICE_FILTER_VALUE_NOT_EMITTED"
    ]
    assert "'south'" in issues[0].message


def test_codebook_validation_reports_uncorrelated_choice_filter():
    xlsx = _codebook_workbook(
        [
            ("select_one regions", "region", ""),
            ("select_one districts", "district", ""),
            (
                "select_one cities",
                "city",
                "concat(${region}, ${district}) = concat(region_code, district_code)",
            ),
        ],
        [
            ("regions", "north", "", "", "North"),
            ("districts", "central", "", "", "Central"),
            ("cities", "city_1", "north", "central", "City 1"),
        ],
        survey_columns=["type", "name", "choice_filter"],
        choice_columns=[
            "list_name",
            "name",
            "region_code",
            "district_code",
            "label",
        ],
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_CHOICE_FILTER_UNCORRELATED"]


def test_codebook_validation_skips_external_choice_filter_schema():
    xlsx = _codebook_workbook(
        [("select_one_from_file cities.csv", "city", "unknown=${missing}")],
        survey_columns=["type", "name", "choice_filter"],
    )

    assert validate_codebook_integrity(xlsx) == []


@pytest.mark.parametrize(
    "expression_column",
    [
        "relevant",
        "constraint",
        "calculation",
        "repeat_count",
        "required",
        "default",
        "read_only",
        "disabled",
        "parameters",
        "label",
        "hint",
    ],
)
def test_codebook_validation_reports_general_reference_case_mismatch(
    expression_column,
):
    xlsx = _codebook_workbook(
        [
            ("integer", "Age", ""),
            ("calculate", "result", "${age}"),
        ],
        survey_columns=["type", "name", expression_column],
    )

    issues = validate_codebook_integrity(xlsx)

    assert [issue.code for issue in issues] == ["CODEBOOK_REFERENCE_CASE_MISMATCH"]
    assert issues[0].field == expression_column
    assert issues[0].owner == {
        "model": "question",
        "name": "result",
        "type": "calculate",
    }
    assert "references 'age'" in issues[0].message
    assert "available exact name: 'Age'" in issues[0].message


def test_codebook_validation_accepts_exact_case_reference():
    xlsx = _codebook_workbook(
        [("integer", "Age", ""), ("calculate", "result", "${Age}")],
        survey_columns=["type", "name", "calculation"],
    )

    assert validate_codebook_integrity(xlsx) == []


def test_general_case_validation_leaves_missing_reference_for_scope_validator():
    xlsx = _codebook_workbook(
        [("calculate", "result", "${not_emitted}")],
        survey_columns=["type", "name", "calculation"],
    )

    assert validate_codebook_integrity(xlsx) == []


def test_composition_errors_block_pyxform_conversion():
    artifact = GeneratedSurveyArtifact(
        _codebook_workbook([("select_one foods", "preferred_food")])
    )

    result = validate_generated_artifact(artifact, converter_cls=ConversionMustNotRun)

    assert result.valid is False
    assert [issue.code for issue in result.errors] == [
        "CODEBOOK_CHOICE_LIST_NOT_EMITTED"
    ]


def test_generated_survey_reports_internal_select_without_choice_group(
    submodule_1,
    root_question_2,
):
    root_question_2.choices = None
    root_question_2.save(update_fields=["choices"])
    form = XLSForm(
        name="Missing choice group",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    result = validate_generated_artifact(
        build_generated_artifact(form), converter_cls=ConversionMustNotRun
    )

    assert [issue.code for issue in result.errors] == ["CODEBOOK_CHOICE_LIST_MISSING"]
    assert result.errors[0].owner == {
        "model": "question",
        "name": root_question_2.name,
    }


def test_generated_survey_reports_choice_group_without_active_choices(
    submodule_1,
    root_question_2,
    choices_1,
):
    choices_1.choices.update(is_active=False)
    form = XLSForm(
        name="Inactive choices",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    result = validate_generated_artifact(
        build_generated_artifact(form), converter_cls=ConversionMustNotRun
    )

    assert [issue.code for issue in result.errors] == [
        "CODEBOOK_CHOICE_LIST_NOT_EMITTED"
    ]
    assert f"choice list '{choices_1.name}'" in result.errors[0].message


def test_generated_survey_reports_blank_choice_value(
    submodule_1,
    root_question_2,
    choices_1,
):
    choice = choices_1.choices.first()
    choices_1.choices.filter(id=choice.id).update(name="   ")
    form = XLSForm(
        name="Blank choice value",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    result = validate_generated_artifact(
        build_generated_artifact(form), converter_cls=ConversionMustNotRun
    )

    assert [issue.code for issue in result.errors] == ["CODEBOOK_CHOICE_VALUE_MISSING"]
    assert result.errors[0].owner == {
        "model": "choice_list",
        "name": choices_1.name,
    }


def test_generated_survey_reports_blank_english_choice_label(
    submodule_1,
    root_question_2,
    choices_1,
):
    choice = choices_1.choices.first()
    choices_1.choices.filter(id=choice.id).update(label="   ")
    form = XLSForm(
        name="Blank choice label",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
        languages=["en"],
    )

    result = validate_generated_artifact(
        build_generated_artifact(form), converter_cls=ConversionMustNotRun
    )

    assert [issue.code for issue in result.errors] == ["CODEBOOK_CHOICE_LABEL_MISSING"]
    assert result.errors[0].owner == {
        "model": "choice",
        "name": choice.name,
        "choice_list": choices_1.name,
    }


def test_generated_survey_reports_invalid_choice_list_name(
    submodule_1,
    root_question_2,
    choices_1,
):
    choices_1.name = "1invalid"
    choices_1.save(update_fields=["name"])
    form = XLSForm(
        name="Invalid choice list name",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    result = validate_generated_artifact(
        build_generated_artifact(form), converter_cls=ConversionMustNotRun
    )

    assert [issue.code for issue in result.errors] == [
        "CODEBOOK_CHOICE_LIST_NAME_INVALID"
    ]
    assert result.errors[0].owner == {
        "model": "choice_list",
        "name": "1invalid",
    }


def test_generated_select_multiple_reports_choice_value_with_whitespace(
    submodule_1,
    root_question_2,
    choices_1,
):
    choice = choices_1.choices.first()
    choices_1.choices.filter(id=choice.id).update(name="with space")
    form = XLSForm(
        name="Invalid choice value",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    result = validate_generated_artifact(
        build_generated_artifact(form), converter_cls=ConversionMustNotRun
    )

    assert [issue.code for issue in result.errors] == ["CODEBOOK_CHOICE_VALUE_INVALID"]
    assert result.errors[0].owner == {
        "model": "choice",
        "name": "with space",
        "choice_list": choices_1.name,
    }


def test_generated_survey_reports_missing_question_name(
    submodule_1,
    root_question_1,
):
    root_question_1.name = ""
    root_question_1.save(update_fields=["name"])
    form = XLSForm(
        name="Missing generated name",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    result = validate_generated_artifact(
        build_generated_artifact(form), converter_cls=ConversionMustNotRun
    )

    assert [issue.code for issue in result.errors] == [
        "CODEBOOK_GENERATED_NAME_MISSING"
    ]
    assert result.errors[0].owner == {"model": "question", "type": "integer"}


def test_generated_survey_reports_invalid_question_name(
    submodule_1,
    root_question_1,
):
    root_question_1.name = "1invalid"
    root_question_1.save(update_fields=["name"])
    form = XLSForm(
        name="Invalid generated name",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    result = validate_generated_artifact(
        build_generated_artifact(form), converter_cls=ConversionMustNotRun
    )

    assert [issue.code for issue in result.errors] == [
        "CODEBOOK_GENERATED_NAME_INVALID"
    ]
    assert result.errors[0].owner == {
        "model": "question",
        "name": "1invalid",
        "type": "integer",
    }


def test_generated_survey_reports_root_and_subquestion_name_collision(
    submodule_1,
    root_question_1,
    sub_question_1,
):
    type(sub_question_1).objects.filter(id=sub_question_1.id).update(
        name=root_question_1.name
    )
    sub_question_1.refresh_from_db()
    form = XLSForm(
        name="Duplicate generated name",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[sub_question_1.id],
        submodules_order=[submodule_1.id],
    )

    result = validate_generated_artifact(
        build_generated_artifact(form), converter_cls=ConversionMustNotRun
    )

    assert [issue.code for issue in result.errors] == [
        "CODEBOOK_GENERATED_NAME_DUPLICATE"
    ]
    assert result.errors[0].owner["model"] == "question"
    assert result.errors[0].owner["name"] == root_question_1.name
    assert result.errors[0].owner["first_model"] == "question"
    assert result.errors[0].owner["first_row"] < result.errors[0].row


def test_generated_survey_reports_invalid_nested_suffix_relationship(
    submodule_1,
    sub_question_4,
):
    sub_question_4.suffix.nested_suffixes.remove(sub_question_4.suffix_2)
    form = XLSForm(
        name="Invalid nested suffix",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[sub_question_4.id],
        submodules_order=[submodule_1.id],
    )

    result = validate_generated_artifact(
        build_generated_artifact(form), converter_cls=ConversionMustNotRun
    )

    assert [issue.code for issue in result.errors] == ["CODEBOOK_NESTED_SUFFIX_INVALID"]
    assert result.errors[0].owner == {
        "model": "SubQuestion",
        "id": sub_question_4.id,
        "name": sub_question_4.name,
    }


def test_generated_survey_accepts_valid_nested_suffix_relationship(
    submodule_1,
    sub_question_4,
):
    sub_question_4.suffix.nested_suffixes.add(sub_question_4.suffix_2)
    form = XLSForm(
        name="Valid nested suffix",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[sub_question_4.id],
        submodules_order=[submodule_1.id],
    )

    artifact = build_generated_artifact(form)

    assert artifact.row_source_map
    assert (
        validate_codebook_integrity(artifact.xlsx_bytes, artifact.row_source_map) == []
    )


def test_generated_survey_keeps_automatic_other_suffix_behavior(
    submodule_1,
    sub_question_1,
):
    sub_question_1.suffix.name = "_oth"
    sub_question_1.suffix.save(update_fields=["name"])
    sub_question_1.save()
    form = XLSForm(
        name="Automatic other suffix",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    artifact = build_generated_artifact(form)
    emitted_sources = list(artifact.row_source_map.values())

    assert any(
        source.get("model") == "SubQuestion" and source.get("id") == sub_question_1.id
        for source in emitted_sources
    )
    assert (
        validate_codebook_integrity(artifact.xlsx_bytes, artifact.row_source_map) == []
    )


def test_generated_survey_reports_recall_period_without_name(
    submodule_1,
    sub_question_2,
    recall_period_1,
):
    type(recall_period_1).objects.filter(id=recall_period_1.id).update(name="")
    form = XLSForm(
        name="Missing recall-period name",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[sub_question_2.id],
        submodules_order=[submodule_1.id],
    )

    result = validate_generated_artifact(
        build_generated_artifact(form), converter_cls=ConversionMustNotRun
    )

    assert [issue.code for issue in result.errors] == [
        "CODEBOOK_RECALL_PERIOD_NAME_MISSING"
    ]
    assert result.errors[0].owner == {
        "model": "SubQuestion",
        "id": sub_question_2.id,
        "name": sub_question_2.name,
    }


def test_generated_survey_accepts_valid_choice_filter(
    submodule_1,
    root_question_1,
    root_question_2,
    choices_1,
    choices_2,
):
    _configure_generated_choice_filter(
        root_question_1, root_question_2, choices_2, choices_1
    )
    form = XLSForm(
        name="Valid choice filter",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )
    artifact = build_generated_artifact(form)

    result = validate_generated_artifact(artifact)

    assert result.valid is True
    assert result.errors == ()


def test_generated_survey_reports_choice_filter_without_configured_source_list(
    submodule_1,
    root_question_1,
    root_question_2,
    choices_1,
    choices_2,
):
    _configure_generated_choice_filter(
        root_question_1, root_question_2, choices_2, choices_1
    )
    choices_1.choice_filter_list = None
    choices_1.save(update_fields=["choice_filter_list"])
    form = XLSForm(
        name="Missing filter source list",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    result = validate_generated_artifact(
        build_generated_artifact(form), converter_cls=ConversionMustNotRun
    )

    assert [issue.code for issue in result.errors] == [
        "CODEBOOK_CHOICE_FILTER_SOURCE_LIST_MISSING"
    ]
    assert result.errors[0].owner == {
        "model": "RootQuestion",
        "id": root_question_2.id,
        "name": root_question_2.name,
    }


def test_generated_survey_reports_incompatible_choice_filter_source_list(
    submodule_1,
    root_question_1,
    root_question_2,
    choices_1,
    choices_2,
):
    _configure_generated_choice_filter(
        root_question_1, root_question_2, choices_2, choices_1
    )
    choices_1.choice_filter_list = ChoiceGroup.objects.create(name="OtherSourceList")
    choices_1.save(update_fields=["choice_filter_list"])
    form = XLSForm(
        name="Incompatible filter source list",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    result = validate_generated_artifact(
        build_generated_artifact(form), converter_cls=ConversionMustNotRun
    )

    assert [issue.code for issue in result.errors] == [
        "CODEBOOK_CHOICE_FILTER_SOURCE_LIST_INCOMPATIBLE"
    ]


def test_generated_survey_reports_unavailable_choice_filter_value(
    submodule_1,
    root_question_1,
    root_question_2,
    choices_1,
    choices_2,
):
    _configure_generated_choice_filter(
        root_question_1, root_question_2, choices_2, choices_1
    )
    source_choice = choices_2.choices.order_by("order", "id").first()
    source_choice.is_active = False
    source_choice.save(update_fields=["is_active"])
    form = XLSForm(
        name="Unavailable filter value",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    result = validate_generated_artifact(
        build_generated_artifact(form), converter_cls=ConversionMustNotRun
    )

    assert [issue.code for issue in result.errors] == [
        "CODEBOOK_CHOICE_FILTER_VALUE_NOT_EMITTED"
    ]
    assert f"'{source_choice.name}'" in result.errors[0].message


def test_generated_survey_reports_reference_case_mismatch_with_question_owner(
    submodule_1,
    root_question_1,
    root_question_2,
):
    root_question_2.relevant = f"${{{root_question_1.name.lower()}}} > 0"
    root_question_2.save(update_fields=["relevant"])
    form = XLSForm(
        name="Question reference case mismatch",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    result = validate_generated_artifact(
        build_generated_artifact(form), converter_cls=ConversionMustNotRun
    )

    assert [issue.code for issue in result.errors] == [
        "CODEBOOK_REFERENCE_CASE_MISMATCH"
    ]
    assert result.errors[0].owner == {
        "model": "RootQuestion",
        "id": root_question_2.id,
        "name": root_question_2.name,
    }
    assert result.errors[0].field == "relevant"


def test_generated_survey_reports_reference_case_mismatch_with_module_owner(
    submodule_1,
    root_question_1,
):
    module = submodule_1.module
    module.relevant = f"${{{root_question_1.name.lower()}}} > 0"
    module.save(update_fields=["relevant"])
    form = XLSForm(
        name="Module reference case mismatch",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    result = validate_generated_artifact(
        build_generated_artifact(form), converter_cls=ConversionMustNotRun
    )

    assert [issue.code for issue in result.errors] == [
        "CODEBOOK_REFERENCE_CASE_MISMATCH"
    ]
    assert result.errors[0].owner == {
        "model": "Module",
        "id": module.id,
        "name": module.name,
    }
    assert result.errors[0].field == "relevant"


def test_empty_external_file_is_a_structured_input_error():
    with pytest.raises(ArtifactInputError) as raised:
        materialize_external_files(
            {"choices.csv": ContentFile(b"", name="choices.csv")}
        )

    assert isinstance(raised.value.issue, ValidationIssue)
    assert raised.value.issue.code == "EXTERNAL_FILE_EMPTY"


def test_xml_conversion_disables_java_validation(monkeypatch):
    calls = {}

    class Converted:
        xform = "<xform/>"

    def convert(xlsform, **kwargs):
        calls.update(kwargs)
        return Converted()

    monkeypatch.setattr("questions.services.xml_conversion.xls2xform.convert", convert)

    conversion = XMLConversion(io.BytesIO(b"xlsx"))
    assert conversion.run() == "<xform/>"
    assert calls["validate"] is False
    assert calls["pretty_print"] is True
    assert calls["enketo"] is False

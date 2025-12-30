from io import BytesIO

from modules.models import Indicator, Module, Submodule
from openpyxl import load_workbook
from questions.models import BaseQuestion, Choice
from questions.services import QuestionsExport


def _choices_rows_for_list(wb_bytes: bytes, list_name: str):
    """Return choices sheet rows for a given list_name, preserving row order in the sheet."""
    wb = load_wb(wb_bytes)
    ws = wb["choices"]
    header = [c.value for c in ws[1]]
    idx_list = header.index("choice_list")
    idx_name = header.index("name")
    # include all label::* columns too (handy for assertions)
    label_cols = [
        i
        for i, h in enumerate(header)
        if isinstance(h, str) and h.startswith("label::")
    ]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx_list] == list_name:
            rows.append(
                {
                    "name": row[idx_name],
                    "labels": {header[i]: row[i] for i in label_cols},
                    "raw": row,
                }
            )
    return rows


def load_wb(b: bytes):
    return load_workbook(BytesIO(b))


def test_questions_export(
    submodule_1,
    root_question_1,
    indicator_1,
    xls_form_data,
):
    export = QuestionsExport()
    assert export.generate_from_submodules(Submodule.objects.all())
    export = QuestionsExport()
    assert export.generate_from_questions(BaseQuestion.objects.all())
    export = QuestionsExport()
    assert export.generate_from_modules(Module.objects.all())
    export = QuestionsExport()
    assert export.generate_from_indicators(Indicator.objects.all())


def test_choices_order_respects_order_then_id(db, question_with_choices):
    """
    Exported choices sheet must list rows for the list in order of Choice.order,
    and when order ties, by id (ascending).
    """
    # Add two extra with same order to verify id tie-break
    cg = question_with_choices["choice_group"]
    # Force same order, different ids (creation order = id ascending)
    Choice.objects.create(
        choice_group=cg, name="42A", label="Tie A", order=10, is_active=True
    )
    Choice.objects.create(
        choice_group=cg, name="42B", label="Tie B", order=10, is_active=True
    )

    export = QuestionsExport(languages=[("en", "English")])
    # Minimal path: generate from questions so exporter computes languages/choice groups, etc.
    wb_bytes = export.generate_from_questions(
        BaseQuestion.objects.filter(id=question_with_choices["base_question"].id)
    )

    rows = _choices_rows_for_list(wb_bytes, list_name=cg.name)
    # Assert we have all choices (original 4 + 2 tie = 6)
    assert [r["name"] for r in rows] == ["1", "0", "777", "888", "42A", "42B"]
    # Confirm English label column is present and aligned
    en_col = "label::English (en)"
    assert all(en_col in r["labels"] for r in rows)
    assert rows[0]["labels"][en_col] == "Yes"
    assert rows[1]["labels"][en_col] == "No"


def test_export_includes_only_active_choices_and_in_correct_order(
    db, question_with_choices
):
    cg = question_with_choices["choice_group"]
    # Make one inactive; it should be excluded from choices sheet
    Choice.objects.filter(choice_group=cg, name="888").update(is_active=False)

    export = QuestionsExport(languages=[("en", "English")])
    wb_bytes = export.generate_from_questions(
        BaseQuestion.objects.filter(id=question_with_choices["base_question"].id)
    )

    rows = _choices_rows_for_list(wb_bytes, list_name=cg.name)
    # 888 removed; order of remaining preserved
    assert [r["name"] for r in rows] == ["1", "0", "777"]

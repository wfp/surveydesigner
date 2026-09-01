from io import BytesIO

from accounts.factories import AdminFactory
from django.contrib.auth import get_user_model
from django.test import TestCase
from modules.factories import ModuleFactory, SubmoduleFactory
from modules.models import Module, Submodule
from openpyxl import Workbook, load_workbook
from organization.models import Organization
from organization.tests.factories import OrganizationFactory
from questions.const import QuestionType
from questions.models import RecallPeriod, RepeatSection, RootQuestion, Suffix
from questions.services import DataImport
from questions.services.questions_import.recall_period import RecallPeriodImport
from questions.services.questions_import.submodule_required_group import (
    SubmoduleRequiredGroupImport,
)
from questions.services.workbook import save_virtual_workbook


def test_sunny_scenario(admin):
    data_import = DataImport(
        "./survey_designer/apps/questions/tests/files/questions.xlsx",
        admin,
        Organization.objects.filter(id=admin.organization.id),
    )
    data_import.process()
    assert data_import.is_valid()
    (
        created,
        created_choices,
        created_suffixes,
        created_recall_periods,
        created_groups,
        created_indicators,
    ) = data_import.create()
    assert created == 10
    assert RepeatSection.objects.count() == 1
    assert len(created_choices) == 2
    assert len(created_suffixes) == 2
    assert len(created_recall_periods) == 2
    assert RootQuestion.objects.count() == 7
    assert RootQuestion.objects.filter(type="calculate").count() == 1
    assert RootQuestion.objects.exclude(calculation="").count() == 2
    rq_with_constraint = RootQuestion.objects.filter(constraint_message__gt="").first()
    assert rq_with_constraint is not None
    assert rq_with_constraint.constraint_message == "test constraint message"
    assert RootQuestion.objects.exclude(parameters="").count() == 2


def test_sunny_scenario_reuses_existing_module_and_submodule_for_case_variant_names(
    setup_upload_base,
):
    _, admin_unicef, _ = setup_upload_base
    workbook = load_workbook(
        "./survey_designer/apps/questions/static/codebook_upload_template.xlsx"
    )

    survey_ws = workbook["survey"]
    question_name = "case_insensitive_import_question"
    survey_ws.cell(row=2, column=1, value="UNICEF_MODULE")
    survey_ws.cell(row=2, column=2, value="UNICEF Module")
    survey_ws.cell(row=2, column=3, value="UNICEF_SUBMODULE_1")
    survey_ws.cell(row=2, column=4, value="UNICEF Submodule 1")
    survey_ws.cell(row=2, column=5, value="text")
    survey_ws.cell(row=2, column=7, value=question_name)
    survey_ws.cell(row=2, column=8, value="Case insensitive import question")
    survey_ws.cell(row=2, column=19, value="description")

    indicators_ws = workbook["indicators"]
    indicators_ws.cell(row=2, column=1, value="test_indicator")
    indicators_ws.cell(row=2, column=2, value="Test Indicator")
    indicators_ws.cell(row=2, column=3, value=question_name)

    module_count_before = Module.objects.filter(name__iexact="unicef_module").count()
    existing_submodule = Submodule.objects.get(name__iexact="unicef_submodule_1")

    data_import = DataImport(
        BytesIO(save_virtual_workbook(workbook)),
        admin_unicef,
        Organization.objects.filter(name="UNICEF_"),
    )
    data_import.process()

    assert data_import.is_valid()

    data_import.create()

    imported_question = RootQuestion.objects.get(name=question_name)
    assert (
        Module.objects.filter(name__iexact="unicef_module").count()
        == module_count_before
    )
    assert Submodule.objects.filter(module__name__iexact="unicef_module").count() == 2
    assert imported_question.submodule.get().id == existing_submodule.id


class TestBulkUploadCaseInsensitiveLookups(TestCase):
    def setUp(self):
        self.organization = OrganizationFactory(name="UNICEF_")
        self.admin = AdminFactory(
            email="admin@unicef.test",
            organization=self.organization,
        )
        self.module = ModuleFactory(name="unicef_module", label="unicef_module")
        self.module.organizations.set([self.organization])
        self.submodule_1 = SubmoduleFactory(
            name="unicef_submodule_1",
            label="unicef_submodule_1",
            module=self.module,
        )
        SubmoduleFactory(
            name="unicef_submodule_2",
            label="unicef_submodule_2",
            module=self.module,
        )

    def test_reuses_existing_module_and_submodule_for_case_variant_names(self):
        workbook = load_workbook(
            "./survey_designer/apps/questions/static/codebook_upload_template.xlsx"
        )

        survey_ws = workbook["survey"]
        question_name = "case_insensitive_import_question"
        survey_ws.cell(row=2, column=1, value="UNICEF_MODULE")
        survey_ws.cell(row=2, column=2, value="UNICEF Module")
        survey_ws.cell(row=2, column=3, value="UNICEF_SUBMODULE_1")
        survey_ws.cell(row=2, column=4, value="UNICEF Submodule 1")
        survey_ws.cell(row=2, column=5, value="text")
        survey_ws.cell(row=2, column=7, value=question_name)
        survey_ws.cell(row=2, column=8, value="Case insensitive import question")
        survey_ws.cell(row=2, column=19, value="description")

        indicators_ws = workbook["indicators"]
        indicators_ws.cell(row=2, column=1, value="test_indicator")
        indicators_ws.cell(row=2, column=2, value="Test Indicator")
        indicators_ws.cell(row=2, column=3, value=question_name)

        module_count_before = Module.objects.filter(
            name__iexact="unicef_module"
        ).count()

        data_import = DataImport(
            BytesIO(save_virtual_workbook(workbook)),
            self.admin,
            Organization.objects.filter(name="UNICEF_"),
        )
        data_import.process()

        self.assertTrue(data_import.is_valid())

        data_import.create()

        imported_question = RootQuestion.objects.get(name=question_name)
        self.assertEqual(
            Module.objects.filter(name__iexact="unicef_module").count(),
            module_count_before,
        )
        self.assertEqual(
            Submodule.objects.filter(module__name__iexact="unicef_module").count(),
            2,
        )
        self.assertEqual(imported_question.submodule.get().id, self.submodule_1.id)


class TestCaseInsensitiveBulkUploadComponentLookups(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            email="bulk-components@example.com",
            password="test_user",
            is_superuser=True,
            is_staff=True,
        )
        self.organization = OrganizationFactory(name="BULK_ORG")
        self.module = ModuleFactory(name="test_module", label="test_module")
        self.module.organizations.set([self.organization])
        self.submodule = SubmoduleFactory(
            name="test_submodule",
            label="test_submodule",
            module=self.module,
        )
        self.suffix = Suffix.objects.create(
            name="Suffix1",
            description="Suffix 1",
            type=QuestionType.TEXT,
        )
        self.recall_period = RecallPeriod.objects.create(
            name="RecallPeriod1",
            description="Recall Period 1",
        )

    def test_recall_period_import_reuses_existing_case_variant_name(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "recall_periods"
        ws.append(["name", "description"])
        ws.append(["recallperiod1", "Updated description"])

        recall_import = RecallPeriodImport(
            ws, self.user, Organization.objects.filter(id=self.organization.id)
        )
        recall_import.process()

        self.assertTrue(recall_import.is_valid())

        created = recall_import.create()

        self.assertEqual(
            RecallPeriod.objects.filter(name__iexact="recallperiod1").count(), 1
        )
        self.assertEqual(created["recallperiod1"].id, self.recall_period.id)

    def test_required_group_import_resolves_case_variant_existing_names(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "required_groups"
        ws.append(["submodule_name", "suffix1", "suffix2", "recall_period"])
        ws.append(["TEST_SUBMODULE", "SUFFIX1", "", "RECALLPERIOD1"])

        required_group_import = SubmoduleRequiredGroupImport(ws, self.user)
        required_group_import.process()

        self.assertTrue(required_group_import.is_valid())

        created = required_group_import.create()

        self.assertEqual(len(created), 1)
        group = self.submodule.required_groups.get()
        self.assertEqual(group.submodule_id, self.submodule.id)
        self.assertEqual(group.required_suffix_id, self.suffix.id)
        self.assertEqual(group.required_recall_period_id, self.recall_period.id)

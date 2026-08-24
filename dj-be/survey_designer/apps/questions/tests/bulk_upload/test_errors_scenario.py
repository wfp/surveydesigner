from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from organization.models import Organization
from questions.services import DataImport
from questions.services.workbook import save_virtual_workbook

_TEST_PARAMETER_NAMES = "upload_organizations_names, module_name, submodule_name, question_name, errors_count, errors_dict"

_TEST_PARAMETERS = (
    (
        # Valid use case
        ["UNICEF_"],
        "unicef_module",
        "unicef_submodule_1",
        "unicef_module_and_submodule_no_errors",
        0,
        None,
    ),
    (
        # Admin referring different organization
        ["UNICEF_"],
        "wfp_module",
        "wfp_submodule_1",
        "unicef_admin_referring_different_organization_error",
        1,
        {
            "Questions Spreadsheet": {
                "errors": {},
                "non_form_errors": [
                    "User has no permission to upload content to module (wfp_module), because module belongs to other organization, or more than one organization."
                ],
            }
        },
    ),
    (
        # Admin referring mismatched module-submodule pair and foreign module
        ["UNICEF_"],
        "wfp_module",
        "unicef_submodule_1",
        "unicef_admin_referring_submodule_not_in_module",
        1,
        {
            "Indicators Spreadsheet": {
                "errors": {},
                "non_form_errors": [
                    "Question: unicef_admin_referring_submodule_not_in_module is not in the system"
                ],
            },
            "Questions Spreadsheet": {
                "errors": {
                    2: {
                        "submodule_name": [
                            "Submodule (unicef_submodule_1) already exists and does not belong to the wfp_module module."
                        ]
                    }
                },
                "non_form_errors": [
                    "User has no permission to upload content to module (wfp_module), because module belongs to other organization, or more than one organization."
                ],
            },
        },
    ),
    (
        # Admin referring mismatched module-submodule pair, submodule existing in different organizations module
        ["UNICEF_"],
        "unicef_module",
        "wfp_submodule_1",
        "unicef_admin_referring_submodule_not_in_module",
        1,
        {
            "Indicators Spreadsheet": {
                "errors": {},
                "non_form_errors": [
                    "Question: unicef_admin_referring_submodule_not_in_module is not in the system"
                ],
            },
            "Questions Spreadsheet": {
                "errors": {
                    2: {
                        "submodule_name": [
                            "Submodule (wfp_submodule_1) already exists and does not belong to the unicef_module module."
                        ]
                    }
                },
                "non_form_errors": [],
            },
        },
    ),
    (
        # Admin referring shared module and submodule
        ["UNICEF_"],
        "shared_module",
        "shared_submodule",
        "unicef_admin_referring_shared_module_submodule",
        1,
        {
            "Questions Spreadsheet": {
                "errors": {},
                "non_form_errors": [
                    "User has no permission to upload content to module (shared_module), because module belongs to other organization, or more than one organization."
                ],
            }
        },
    ),
    (
        # Given question exists but is shared, admin tries to add submodule with proper organization.
        ["UNICEF_"],
        "unicef_module",
        "unicef_submodule_2",
        "wfp_unicef_shared_question",
        1,
        {
            "Questions Spreadsheet": {
                "errors": {},
                "non_form_errors": [
                    "Cannot process existing question (wfp_unicef_shared_question), "
                    "because this question's modules belongs to different organizations "
                    "than declared.If you want to add this question to submodule belonging to "
                    "different organizations subset, please ask global admin to do it through CMS interface."
                ],
            }
        },
    ),
    (
        # Given question exists but is shared, admin tries to add submodule with proper organization,
        # edge case with matching org count, where Count and __in filtering fails.
        ["UNICEF_", "WFP_"],
        "unicef_module",
        "unicef_submodule_2",
        "red_cross_unicef_shared_question",
        1,
        {
            "Questions Spreadsheet": {
                "errors": {},
                "non_form_errors": [
                    "User has no permission to upload content to module (unicef_module), "
                    "because module belongs to other organization, or more than one organization.",
                    "Cannot process existing question (red_cross_unicef_shared_question), "
                    "because this question's modules belongs to different organizations "
                    "than declared.If you want to add this question to submodule belonging to "
                    "different organizations subset, please ask global admin to do it through CMS interface.",
                ],
            }
        },
    ),
    (
        # Given question does not exist in the system
        ["UNICEF_"],
        "unicef_module",
        "unicef_submodule_1",
        "unicef_module_and_submodule_different_to_indicator",
        1,
        {
            "Indicators Spreadsheet": {
                "errors": {},
                "non_form_errors": ["Question: non_existent is not in the system"],
            }
        },
    ),
)


@pytest.mark.parametrize(_TEST_PARAMETER_NAMES, _TEST_PARAMETERS)
def test_errors(
    upload_organizations_names,
    setup_upload_base,
    module_name,
    submodule_name,
    question_name,
    errors_count,
    errors_dict,
):
    _, admin_unicef, _ = setup_upload_base
    expecting_errors = errors_count != 0

    test_workbook: Workbook = load_workbook(
        "./survey_designer/apps/questions/static/codebook_upload_template.xlsx"
    )

    survey_ws: Worksheet = test_workbook["survey"]

    # Column values referring codebook_upload_template.xlsx.
    survey_ws.cell(row=2, column=1, value=module_name)
    survey_ws.cell(row=2, column=2, value=module_name)
    survey_ws.cell(row=2, column=3, value=submodule_name)
    survey_ws.cell(row=2, column=4, value=submodule_name)
    survey_ws.cell(row=2, column=7, value=question_name)
    survey_ws.cell(row=2, column=8, value=question_name)

    survey_ws.cell(row=2, column=5, value="text")
    survey_ws.cell(row=2, column=9, value="test")
    survey_ws.cell(row=2, column=19, value="description")

    # Fill in the "indicators" sheet
    indicators_ws: Worksheet = test_workbook["indicators"]
    indicators_ws.cell(row=2, column=1, value="test_indicator")
    indicators_ws.cell(row=2, column=2, value="Test Indicator")
    if question_name == "unicef_module_and_submodule_different_to_indicator":
        question_name = "non_existent"
    indicators_ws.cell(row=2, column=3, value=question_name)

    bytes_test_workbook = BytesIO(save_virtual_workbook(test_workbook))

    data_import = DataImport(
        bytes_test_workbook,
        admin_unicef,
        Organization.objects.filter(name__in=upload_organizations_names),
    )
    data_import.process()

    assert data_import.is_valid() != expecting_errors

    if not expecting_errors:
        return

    errors = data_import.get_errors()
    assert errors == errors_dict

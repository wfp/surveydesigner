# tests/test_choices_import.py
# tests/utils_workbook.py
from io import BytesIO

import pytest
from django.conf import settings
from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from questions.models import Choice, ChoiceGroup
from questions.services.questions_import.choices import ChoicesImport


def build_choices_wb(choices_rows, languages=("en", "English"), with_other_sheets=True):
    """
    choices_rows: list of dicts like:
      {"choice_list": "Yesnodkref", "name": "1", "labels": {"en": "Yes"}}
    Returns BytesIO of an xlsx that includes the required sheets for DataImport.
    """
    wb = Workbook()
    ws_survey = wb.active
    ws_survey.title = "survey"
    ws_survey.append(
        [
            "module_name",
            "module_label",
            "submodule_name",
            "submodule_label",
            "type",
            "choice_list",
            "name",
            "label::English (en)",
            "hint",
            "suffix1",
            "suffix2",
            "recall_period",
            "relevant",
            "constraint",
            "constraint_message",
            "calculation",
            "description",
            "repeat",
            "repeat_count",
            "appearance",
            "parameters",
            "default",
            "disabled",
            "required",
            "read_only",
        ]
    )

    # choices
    ws_choices = wb.create_sheet("choices")
    ws_choices.append(["choice_list", "name", "label::English (en)"])
    for row in choices_rows:
        label = row["labels"].get("en") or ""
        ws_choices.append([row["choice_list"], row["name"], label])

    if with_other_sheets:
        wb.create_sheet("suffixes").append(
            ["name", "description", "type", "choicelist", "suffix"]
        )
        wb.create_sheet("recall_periods").append(["name", "description"])
        wb.create_sheet("required_groups").append(
            ["submodule_name", "suffix1", "suffix2", "recall_period"]
        )
        ws_ind = wb.create_sheet("indicators")
        ws_ind.append(
            ["indicator_name", "indicator_label", "question_name"]
        )  # keep empty body

    # return as BytesIO
    return BytesIO(save_virtual_workbook(wb))


@pytest.mark.django_db
def test_import_sets_order_from_row_index_for_new_choices(admin):
    rows = [
        {"choice_list": "Yesnodkref", "name": "1", "labels": {"en": "Yes"}},
        {"choice_list": "Yesnodkref", "name": "0", "labels": {"en": "No"}},
        {
            "choice_list": "Yesnodkref",
            "name": "777",
            "labels": {"en": "Prefer not to answer"},
        },
        {"choice_list": "Yesnodkref", "name": "888", "labels": {"en": "Don't Know"}},
    ]
    xlsx = build_choices_wb(rows, with_other_sheets=False)

    wb = load_workbook(xlsx)
    ci = ChoicesImport(
        choices_sheet=wb["choices"],
        languages_dict=dict(settings.LANGUAGES),
        user=admin,
    )
    ci.process()
    assert ci.is_valid()

    created = ci.create(skip_saving=False)
    cg = created["Yesnodkref"]
    names_orders = list(
        Choice.objects.filter(choice_group=cg)
        .order_by("order", "id")
        .values_list("name", "order")
    )
    assert names_orders == [("1", 1), ("0", 2), ("777", 3), ("888", 4)]


@pytest.mark.django_db
def test_reimport_does_not_change_existing_nonzero_orders(admin):
    # Initial import
    rows = [
        {"choice_list": "Yesnodkref", "name": "1", "labels": {"en": "Yes"}},
        {"choice_list": "Yesnodkref", "name": "0", "labels": {"en": "No"}},
        {
            "choice_list": "Yesnodkref",
            "name": "777",
            "labels": {"en": "Prefer not to answer"},
        },
    ]
    wb1 = load_workbook(build_choices_wb(rows, with_other_sheets=False))
    ci1 = ChoicesImport(wb1["choices"], dict(settings.LANGUAGES), admin)
    ci1.process()
    assert ci1.is_valid()
    ci1.create()

    cg = ChoiceGroup.objects.get(name="Yesnodkref")
    # Guard: current orders are 1,2,3
    assert list(
        Choice.objects.filter(choice_group=cg)
        .order_by("order")
        .values_list("name", flat=True)
    ) == ["1", "0", "777"]

    # Re-import with different row order + one new item
    rows_re = [
        {
            "choice_list": "Yesnodkref",
            "name": "777",
            "labels": {"en": "Prefer not to answer"},
        },  # moved to top
        {"choice_list": "Yesnodkref", "name": "1", "labels": {"en": "Yes"}},
        {"choice_list": "Yesnodkref", "name": "0", "labels": {"en": "No"}},
        {
            "choice_list": "Yesnodkref",
            "name": "888",
            "labels": {"en": "Don't Know"},
        },  # new
    ]
    wb2 = load_workbook(build_choices_wb(rows_re, with_other_sheets=False))
    ci2 = ChoicesImport(wb2["choices"], dict(settings.LANGUAGES), admin)
    ci2.process()
    assert ci2.is_valid()
    ci2.create()

    # Existing orders should stay the same; new one gets appended with next index
    names_orders = list(
        Choice.objects.filter(choice_group=cg)
        .order_by("order", "id")
        .values_list("name", "order")
    )
    assert names_orders == [("1", 1), ("0", 2), ("777", 3), ("888", 4)]


@pytest.mark.django_db
def test_reimport_repairs_zero_or_null_order(admin):
    rows = [
        {"choice_list": "Yesnodkref", "name": "A", "labels": {"en": "A"}},
        {"choice_list": "Yesnodkref", "name": "B", "labels": {"en": "B"}},
        {"choice_list": "Yesnodkref", "name": "C", "labels": {"en": "C"}},
    ]
    wb1 = load_workbook(build_choices_wb(rows, with_other_sheets=False))
    ci1 = ChoicesImport(wb1["choices"], dict(settings.LANGUAGES), admin)
    ci1.process()
    assert ci1.is_valid()
    ci1.create()

    cg = ChoiceGroup.objects.get(name="Yesnodkref")
    # Manually break one order to 0
    Choice.objects.filter(choice_group=cg, name="B").update(order=0)

    # Re-import same rows -> "B" should be repaired to 2
    wb2 = load_workbook(build_choices_wb(rows, with_other_sheets=False))
    ci2 = ChoicesImport(wb2["choices"], dict(settings.LANGUAGES), admin)
    ci2.process()
    assert ci2.is_valid()
    ci2.create()

    names_orders = list(
        Choice.objects.filter(choice_group=cg)
        .order_by("order", "id")
        .values_list("name", "order")
    )
    assert names_orders == [("A", 1), ("B", 2), ("C", 3)]


@pytest.mark.django_db
def test_multiple_choice_lists_have_independent_row_order(admin):
    rows = [
        {"choice_list": "Yesno", "name": "1", "labels": {"en": "Yes"}},
        {"choice_list": "Yesno", "name": "0", "labels": {"en": "No"}},
        {"choice_list": "Sex", "name": "0", "labels": {"en": "Female"}},
        {"choice_list": "Sex", "name": "1", "labels": {"en": "Male"}},
    ]
    wb = load_workbook(build_choices_wb(rows, with_other_sheets=False))
    ci = ChoicesImport(wb["choices"], dict(settings.LANGUAGES), admin)
    ci.process()
    assert ci.is_valid()
    ci.create()

    yesno = ChoiceGroup.objects.get(name="Yesno")
    sex = ChoiceGroup.objects.get(name="Sex")
    assert list(
        Choice.objects.filter(choice_group=yesno)
        .order_by("order")
        .values_list("name", flat=True)
    ) == ["1", "0"]
    assert list(
        Choice.objects.filter(choice_group=sex)
        .order_by("order")
        .values_list("name", flat=True)
    ) == ["0", "1"]


class TestCaseInsensitiveChoicesImport(TestCase):
    def setUp(self):
        self.admin = get_user_model().objects.create_user(
            email="choices-import@example.com",
            password="test_user",
            is_superuser=True,
            is_staff=True,
        )
        self.choice_group = ChoiceGroup.objects.create(name="YesNoDkRef")

    def test_reuses_existing_choice_group_for_case_variant_name(self):
        rows = [
            {"choice_list": "yesnodkref", "name": "1", "labels": {"en": "Yes"}},
            {"choice_list": "yesnodkref", "name": "0", "labels": {"en": "No"}},
        ]
        wb = load_workbook(build_choices_wb(rows, with_other_sheets=False))
        ci = ChoicesImport(wb["choices"], dict(settings.LANGUAGES), self.admin)
        ci.process()

        self.assertTrue(ci.is_valid())

        created = ci.create()

        self.assertEqual(
            ChoiceGroup.objects.filter(name__iexact="yesnodkref").count(), 1
        )
        self.assertEqual(created["yesnodkref"].id, self.choice_group.id)

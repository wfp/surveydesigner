import datetime
import io
import os
import uuid

import pandas as pd
import pytest
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from openpyxl import load_workbook
from questions.const import QuestionType
from questions.models import ChoiceGroupFile, RootQuestion, SubQuestion, Suffix
from questions.services import DocConversion, XLSForm, XMLConversion


def test_xls_form_generation(
    submodule_1,
    sub_question_1,
    sub_question_2,
    sub_question_3,
    sub_question_4,
    xls_form_data,
    indicator_2,
    submodule_2,
):
    subquestion_ids = [
        sub_question_1.id,
        sub_question_2.id,
        sub_question_3.id,
        sub_question_4.id,
    ]
    xls_form = XLSForm(
        name="test_name",
        submodule_ids=[submodule_1.id],
        sub_question_ids=subquestion_ids,
        submodules_order={
            submodule_1.module.id: [submodule_1.id],
            submodule_2.module.id: [submodule_2.id],
        },
        languages=["en", "fr"],
        indicators=[indicator_2.id],
    )
    assert xls_form.generate()

    xls_form = XLSForm(
        name="test_name",
        submodule_ids=[submodule_1.id],
        sub_question_ids=subquestion_ids,
        submodules_order={submodule_1.module.id: [submodule_1.id]},
    )
    assert xls_form.generate()


def test_xls_form_includes_survey_designer_metadata(submodule_1):
    xls_form = XLSForm(
        name="test_name",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    workbook = load_workbook(io.BytesIO(xls_form.generate()))
    survey_sheet = workbook["survey"]
    headers = [cell.value for cell in survey_sheet[1]]
    rows = {
        row[headers.index("name")]: dict(zip(headers, row))
        for row in survey_sheet.iter_rows(min_row=2, values_only=True)
        if row[headers.index("name")]
    }

    expected_defaults = {
        "sd_metadata_generated_by": "survey_designer",
        "sd_metadata_export_id": xls_form.export_id,
        "sd_metadata_generator_version": xls_form.get_generator_version(),
    }
    for name, default in expected_defaults.items():
        assert rows[name]["type"] == "hidden"
        assert rows[name]["default"] == default

    exported_at = rows["sd_metadata_exported_at"]
    assert exported_at["type"] == "hidden"
    parsed_timestamp = datetime.datetime.fromisoformat(
        exported_at["default"].replace("Z", "+00:00")
    )
    assert parsed_timestamp.tzinfo == datetime.timezone.utc
    assert parsed_timestamp.microsecond == 0


def test_xls_form_settings_include_versioned_form_id(submodule_1, mocker):
    mocker.patch("questions.services.xls_form.SURVEY_DESIGNER_VERSION", "20260716.1")
    token_urlsafe = mocker.patch(
        "questions.services.xls_form.secrets.token_urlsafe",
        return_value="K8s2pQx9aBc",
    )

    xls_form = XLSForm(
        name="test_name",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    workbook = load_workbook(io.BytesIO(xls_form.generate()))
    settings_rows = list(workbook["settings"].iter_rows(values_only=True))
    settings = dict(zip(settings_rows[0], settings_rows[1]))

    assert settings["form_id"] == "surveydesigner_v202607161_K8s2pQx9aBc"
    assert settings["version"] == "_v202607161"
    assert xls_form.export_id == settings["form_id"]
    token_urlsafe.assert_called_once_with(nbytes=8)


@pytest.mark.parametrize(
    ("generator_version", "expected"),
    [
        ("20260716.1", "_v202607161"),
        ("v202607161", "_v202607161"),
        ("", "_vdev"),
    ],
)
def test_get_form_version(mocker, generator_version, expected):
    mocker.patch(
        "questions.services.xls_form.SURVEY_DESIGNER_VERSION", generator_version
    )

    assert XLSForm.get_form_version() == expected


def test_xls_form_rejects_survey_designer_metadata_name_collision(submodule_1):
    reserved_name = "sd_metadata_export_id"
    question = RootQuestion.objects.filter(submodule=submodule_1).first()
    question.name = reserved_name
    question.save()

    xls_form = XLSForm(
        name="test_name",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    with pytest.raises(ValidationError, match=reserved_name):
        xls_form.generate()


def test_xml_conversion(
    submodule_1,
    sub_question_1,
    sub_question_2,
    sub_question_3,
    sub_question_4,
    xls_form_data,
):
    subquestion_ids = [
        sub_question_1.id,
        sub_question_2.id,
        sub_question_3.id,
        sub_question_4.id,
    ]
    xls_form = XLSForm(
        name="test_name",
        submodule_ids=[submodule_1.id],
        sub_question_ids=subquestion_ids,
        submodules_order={submodule_1.module.id: [submodule_1.id]},
        languages=["en", "fr"],
    )
    xlsx_b = io.BytesIO(xls_form.generate())
    xml_conversion = XMLConversion(xlsx_b)
    assert xml_conversion.run()
    for warning in xml_conversion.warnings:
        assert "'disabled' column header" not in warning


def test_doc_conversion(
    submodule_1,
    sub_question_1,
    sub_question_2,
    sub_question_3,
    sub_question_4,
    xls_form_data,
):
    subquestion_ids = [
        sub_question_1.id,
        sub_question_2.id,
        sub_question_3.id,
        sub_question_4.id,
    ]
    languages = ["en", "fr"]
    xls_form = XLSForm(
        name="test_name",
        submodule_ids=[submodule_1.id],
        sub_question_ids=subquestion_ids,
        submodules_order={submodule_1.module.id: [submodule_1.id]},
        languages=languages,
    )
    xlsx_b = io.BytesIO(xls_form.generate())
    doc_conversion = DocConversion(xlsx_b, languages)
    assert doc_conversion.run()


def test_xls_form_collects_external_files_from_choice_group_file(
    tmp_path,
    settings,
    submodule_1,
    root_question_1,
):
    settings.MEDIA_ROOT = tmp_path
    choices_file = ChoiceGroupFile.objects.create(
        name=f"FruitChoices_{uuid.uuid4().hex}",
        csv_file=ContentFile(b"name,color\nbanana,yellow\n", name="fruits.csv"),
    )
    root_question_1.type = QuestionType.SELECT_ONE_FROM_FILE
    root_question_1.choices = None
    root_question_1.choices_file = choices_file
    root_question_1.save()

    xls_form = XLSForm(
        name="test_name",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )
    xls_form.generate()

    expected_name = os.path.basename(choices_file.csv_file.name)
    assert expected_name in xls_form.external_files
    file_field = xls_form.external_files[expected_name]
    assert os.path.basename(file_field.name) == expected_name


def test_xls_form_collects_external_files_from_suffix_choice_group_file(
    tmp_path,
    settings,
    submodule_1,
):
    settings.MEDIA_ROOT = tmp_path
    choices_file = ChoiceGroupFile.objects.create(
        name=f"ChoiceFile_{uuid.uuid4().hex}",
        csv_file=ContentFile(
            b"name,color\nbanana,yellow\n",
            name=f"fruits_{uuid.uuid4().hex}.csv",
        ),
    )
    root_question = RootQuestion.objects.create(
        name=f"RootQuestion_{uuid.uuid4().hex}",
        label="Root Question",
        type=QuestionType.TEXT,
    )
    root_question.submodule.add(submodule_1)

    suffix = Suffix.objects.create(
        name=f"SuffixFile_{uuid.uuid4().hex}",
        type=QuestionType.SELECT_ONE_FROM_FILE,
        choices_file=choices_file,
    )
    sub_question = SubQuestion.objects.create(
        root_question=root_question,
        name=f"SubQuestion_{uuid.uuid4().hex}",
        suffix=suffix,
        label="Sub Question",
    )

    xls_form = XLSForm(
        name="test_name",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[sub_question.id],
        submodules_order=[submodule_1.id],
    )
    xls_form.generate()

    csv_name = os.path.basename(choices_file.csv_file.name)
    assert csv_name in xls_form.external_files


def test_xls_form_includes_module_relevant(submodule_1, root_question_1):
    module = submodule_1.module
    module.relevant = f"${{{root_question_1.name}}} > 0"
    module.save()

    xls_form = XLSForm(
        name="test_name",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[],
        submodules_order=[submodule_1.id],
    )

    xlsx_b = io.BytesIO(xls_form.generate())
    xls = pd.read_excel(xlsx_b, sheet_name="survey")

    module_row = xls.loc[xls["name"] == f"{module.name}_module"].iloc[0]

    assert module_row["relevant"] == module.relevant


def test_xls_form_choice_filtering(
    submodule_1,
    sub_question_1,
    sub_question_2,
    choices_1,
    choices_2,
):
    for sub_question, choices in zip(
        [sub_question_1, sub_question_2], [choices_1, choices_2]
    ):
        root_question = sub_question.root_question
        root_question.type = "select_one"
        root_question.choices = choices
        root_question.save()
        sub_question.save()
    choice_1 = choices_1.choices.get(name="1")
    choice_2 = choices_1.choices.get(name="2")
    choice_3 = choices_2.choices.get(name="3")
    choice_4 = choices_2.choices.get(name="4")

    choice_1.choice_filter_name = choice_3
    choice_2.choice_filter_name = choice_4
    choice_1.save()
    choice_2.save()
    languages = ["en", "fr"]

    # Generate the XLSForm
    xls_form = XLSForm(
        name="test_name",
        submodule_ids=[submodule_1.id],
        sub_question_ids=[sub_question_1.id, sub_question_2.id],
        submodules_order=[submodule_1.id],
        languages=languages,
    )
    xlsx_b = io.BytesIO(xls_form.generate())
    doc_conversion = DocConversion(xlsx_b, languages)
    assert doc_conversion.run()

    # Load the generated XLSForm
    xls = pd.read_excel(xlsx_b, sheet_name="choices")
    # Validate the choice filters
    assert len(xls) == 4, "Expected 4 choices"
    assert (
        xls.loc[xls["name"] == int(choice_1.name), "choice_filter_name"].values[0] == 3
    )
    assert (
        xls.loc[xls["name"] == int(choice_2.name), "choice_filter_name"].values[0] == 4
    )
    # Assert they're showing as nan (empty cell on xls)
    assert pd.isna(
        xls.loc[xls["name"] == int(choice_3.name), "choice_filter_name"].values[0]
    )
    assert pd.isna(
        xls.loc[xls["name"] == int(choice_4.name), "choice_filter_name"].values[0]
    )

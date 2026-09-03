"""Exact XLSForm artifact validation used by publication boundaries.

This module deliberately stops at pyxform's Python conversion boundary.  The
application does not install or invoke Java/ODK Validate.  The small XML
compatibility validator below is versioned so that its coverage can be
expanded without changing the public issue contract.
"""

from __future__ import annotations

import hashlib
import io
import re
from collections import defaultdict
from dataclasses import dataclass, field, replace
from types import MappingProxyType
from typing import Any, Mapping, Sequence
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from pyxform.parsing.expression import is_xml_tag, parse_expression

PYXFORM_VERSION = "4.5.0"
COMPATIBILITY_VERSION = "1.0"
_EMPTY_ARTIFACT_HASH = "sha256:" + hashlib.sha256(b"").hexdigest()
_XFORMS_NS = "http://www.w3.org/2002/xforms"
_XHTML_NS = "http://www.w3.org/1999/xhtml"
_EXTERNAL_REFERENCE = re.compile(r"jr://file-csv/([^/]+)$")
_ENGLISH_LABEL_COLUMN = re.compile(r"^label::.*\(\s*en\s*\)$", re.IGNORECASE)
_INTERNAL_SELECT_TYPES = ("select_one", "select_multiple")
_EXTERNAL_SELECT_TYPES = ("select_one_from_file", "select_multiple_from_file")
_END_SURVEY_ROW_TYPES = frozenset(("end_group", "end_repeat"))
_RESERVED_SURVEY_NAMES = frozenset(("meta",))
_SURVEY_METADATA_TYPES = frozenset(("start", "end", "today", "deviceid"))
_REFERENCE_CASE_IGNORED_COLUMNS = frozenset(("type", "name", "choice_filter"))
_SUPPORTED_GENERATED_TYPES = frozenset(
    (
        "acknowledge",
        "audio",
        "background-audio",
        "barcode",
        "calculate",
        "date",
        "dateTime",
        "decimal",
        "geopoint",
        "geoshape",
        "geotrace",
        "hidden",
        "image",
        "integer",
        "note",
        "rank",
        "select_multiple",
        "select_multiple_from_file",
        "select_one",
        "select_one_from_file",
        "text",
        "time",
        "video",
    )
)


class ArtifactInputError(Exception):
    """The submitted/generated artifact cannot be published."""

    def __init__(self, issue: "ValidationIssue") -> None:
        super().__init__(issue.message)
        self.issue = issue


class ArtifactInfrastructureError(Exception):
    """The generator or artifact storage could not produce a stable artifact."""


class ValidatorInfrastructureError(Exception):
    """The configured validator could not complete reliably."""


@dataclass(frozen=True)
class ValidationIssue:
    code: str
    layer: str
    severity: str
    message: str
    owner: Mapping[str, Any] | None = None
    field: str | None = None
    sheet: str | None = None
    column: str | None = None
    row: int | None = None

    def as_dict(self) -> dict[str, Any]:
        result: dict[str, Any] = {
            "code": self.code,
            "layer": self.layer,
            "severity": self.severity,
            "message": self.message,
        }
        for key in ("owner", "field", "sheet", "column", "row"):
            value = getattr(self, key)
            if value is not None:
                result[key] = dict(value) if key == "owner" else value
        return result

    to_dict = as_dict


def compute_artifact_hash(
    xlsx_bytes: bytes, external_files: Mapping[str, bytes] | None = None
) -> str:
    """Return a stable, unambiguous hash for the complete publication artifact."""

    digest = hashlib.sha256()
    digest.update(b"survey-designer-artifact-v1\0")

    def add_part(value: bytes) -> None:
        digest.update(len(value).to_bytes(8, byteorder="big"))
        digest.update(value)

    add_part(bytes(xlsx_bytes))
    for filename in sorted((external_files or {})):
        name_bytes = str(filename).encode("utf-8")
        add_part(name_bytes)
        add_part(bytes((external_files or {})[filename]))
    return f"sha256:{digest.hexdigest()}"


@dataclass(frozen=True)
class GeneratedSurveyArtifact:
    """The immutable bytes and conversion output for one generated survey."""

    xlsx_bytes: bytes
    external_files: Mapping[str, bytes] = field(default_factory=dict)
    artifact_hash: str | None = None
    xml: str | None = None
    row_source_map: Mapping[Any, Any] | None = None
    form_name: str = "survey"

    def __post_init__(self) -> None:
        xlsx_bytes = bytes(self.xlsx_bytes)
        external_files = {
            str(name): bytes(content)
            for name, content in dict(self.external_files).items()
        }
        object.__setattr__(self, "xlsx_bytes", xlsx_bytes)
        object.__setattr__(self, "external_files", MappingProxyType(external_files))
        if self.row_source_map is not None:
            object.__setattr__(
                self,
                "row_source_map",
                MappingProxyType(dict(self.row_source_map)),
            )
        artifact_hash = self.artifact_hash or compute_artifact_hash(
            xlsx_bytes, external_files
        )
        object.__setattr__(self, "artifact_hash", artifact_hash)


@dataclass(frozen=True)
class ValidationResult:
    """Public validation response plus the successful exact artifact internally."""

    valid: bool
    artifact_hash: str = _EMPTY_ARTIFACT_HASH
    errors: Sequence[ValidationIssue] = field(default_factory=tuple)
    warnings: Sequence[ValidationIssue] = field(default_factory=tuple)
    validator: Mapping[str, str] = field(
        default_factory=lambda: {
            "pyxform": PYXFORM_VERSION,
            "compatibility": COMPATIBILITY_VERSION,
        }
    )
    artifact: GeneratedSurveyArtifact | None = field(default=None, repr=False)

    def __post_init__(self) -> None:
        errors = tuple(self.errors)
        warnings = tuple(self.warnings)
        object.__setattr__(self, "errors", errors)
        object.__setattr__(self, "warnings", warnings)
        object.__setattr__(self, "valid", not errors)
        object.__setattr__(self, "validator", MappingProxyType(dict(self.validator)))

    def as_dict(self) -> dict[str, Any]:
        return {
            "valid": self.valid,
            "artifact_hash": self.artifact_hash,
            "errors": [issue.as_dict() for issue in self.errors],
            "warnings": [issue.as_dict() for issue in self.warnings],
            "validator": {
                "pyxform": self.validator.get("pyxform", PYXFORM_VERSION),
                "compatibility": self.validator.get(
                    "compatibility", COMPATIBILITY_VERSION
                ),
            },
        }

    # This alias makes the contract convenient for serializers and callers that
    # use the usual ``to_dict`` naming.
    to_dict = as_dict


def _issue(
    code: str,
    layer: str,
    message: str,
    *,
    severity: str = "error",
    **details: Any,
) -> ValidationIssue:
    return ValidationIssue(
        code=code,
        layer=layer,
        severity=severity,
        message=str(message),
        **{key: value for key, value in details.items() if value is not None},
    )


def _normalise_messages(
    messages: Sequence[Any], *, warning: bool
) -> list[ValidationIssue]:
    return [
        _issue(
            "PYXFORM_WARNING" if warning else "PYXFORM_CONVERSION_ERROR",
            "pyxform",
            str(message),
            severity="warning" if warning else "error",
        )
        for message in messages
        if str(message).strip()
    ]


def validate_xml_compatibility(
    xml: str | bytes | None, external_files: Mapping[str, bytes] | None = None
) -> list[ValidationIssue]:
    """Check minimal pyxform output and exact-artifact file references.

    This is not a JavaRosa or ODK Validate replacement. Pyxform performs the
    XLSForm checks; this seam only confirms the generated XForm shell and that
    referenced CSV files are present in the materialized artifact.
    """

    try:
        root = ET.fromstring(xml)
    except (ET.ParseError, TypeError, ValueError) as exc:
        return [
            _issue(
                "XML_MALFORMED",
                "compatibility",
                f"Generated XML is not well formed: {exc}",
                sheet="survey",
            )
        ]

    issues: list[ValidationIssue] = []
    if root.tag != f"{{{_XHTML_NS}}}html":
        issues.append(
            _issue(
                "XML_ROOT_INVALID",
                "compatibility",
                "Generated XForm must have an XHTML html root element.",
                sheet="survey",
            )
        )

    model = root.find(f".//{{{_XFORMS_NS}}}model")
    body = root.find(f".//{{{_XHTML_NS}}}body")
    instances = root.findall(f".//{{{_XFORMS_NS}}}instance")
    if model is None:
        issues.append(
            _issue(
                "XML_MODEL_MISSING",
                "compatibility",
                "Generated XForm is missing its xforms model.",
                sheet="survey",
            )
        )
    if body is None:
        issues.append(
            _issue(
                "XML_BODY_MISSING",
                "compatibility",
                "Generated XForm is missing its XHTML body.",
                sheet="survey",
            )
        )
    if not instances:
        issues.append(
            _issue(
                "XML_INSTANCE_MISSING",
                "compatibility",
                "Generated XForm is missing a primary instance.",
                sheet="survey",
            )
        )

    available_files = set((external_files or {}).keys())
    for element in root.iter():
        for attribute in ("src", "href"):
            value = element.get(attribute)
            if not value:
                continue
            match = _EXTERNAL_REFERENCE.match(value)
            if match and match.group(1) not in available_files:
                issues.append(
                    _issue(
                        "EXTERNAL_FILE_MISSING",
                        "compatibility",
                        f"Generated XForm references external file '{match.group(1)}', but that file is not materialized.",
                        sheet="survey",
                        field=match.group(1),
                    )
                )
    return issues


def materialize_external_files(
    external_files: Mapping[str, Any] | None,
) -> dict[str, bytes]:
    """Read every selected external file exactly once and retain its bytes."""

    materialized: dict[str, bytes] = {}
    for filename, file_obj in (external_files or {}).items():
        filename = str(filename)
        try:
            stream = file_obj.open("rb")
            try:
                content = stream.read()
            finally:
                stream.close()
        except FileNotFoundError:
            raise ArtifactInputError(
                _issue(
                    "EXTERNAL_FILE_MISSING",
                    "composition",
                    f"External file '{filename}' could not be found.",
                    field=filename,
                )
            )
        except OSError as exc:
            raise ArtifactInfrastructureError(
                f"Unable to read external file '{filename}': {exc}"
            ) from exc

        if not content:
            raise ArtifactInputError(
                _issue(
                    "EXTERNAL_FILE_EMPTY",
                    "composition",
                    f"External file '{filename}' is empty.",
                    field=filename,
                )
            )
        materialized[filename] = content
    return materialized


def _sheet_rows_by_header(
    worksheet: Any,
) -> tuple[dict[str, int], list[tuple[Any, ...]]]:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}, []

    headers = {
        str(value).strip(): index
        for index, value in enumerate(rows[0])
        if value is not None and str(value).strip()
    }
    return headers, rows[1:]


def _row_value(row: tuple[Any, ...], headers: Mapping[str, int], column: str) -> str:
    index = headers.get(column)
    if index is None or index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _english_label_column(headers: Mapping[str, int]) -> str | None:
    if "label" in headers:
        return "label"
    return next(
        (column for column in headers if _ENGLISH_LABEL_COLUMN.match(column)), None
    )


def _internal_select_declaration(
    row: tuple[Any, ...], survey_headers: Mapping[str, int]
) -> tuple[str, str, str] | None:
    declaration = _row_value(row, survey_headers, "type")
    question_type, _, inline_list_name = declaration.partition(" ")
    if question_type not in _INTERNAL_SELECT_TYPES:
        return None

    list_column = "choice_list" if "choice_list" in survey_headers else "type"
    list_name = (
        _row_value(row, survey_headers, "choice_list")
        if "choice_list" in survey_headers
        else inline_list_name.strip()
    )
    return question_type, list_name, list_column


def _choice_filter_references(
    expression: str,
) -> tuple[set[str], set[str], set[tuple[str, str]]]:
    """Return emitted columns, question references, and direct correlations."""

    tokens = [
        token for token in parse_expression(expression) if token.type != "WHITESPACE"
    ]
    columns: set[str] = set()
    questions: set[str] = set()

    for index, token in enumerate(tokens):
        value = str(token)
        if token.type == "PYXFORM_REF":
            question_name = value[2:-1]
            if not question_name.startswith("last-saved#"):
                questions.add(question_name)
        elif token.type in ("NAME", "XPATH_PRED_START"):
            if index and tokens[index - 1].type == "PATH_SEP":
                continue
            columns.add(value[:-1] if token.type == "XPATH_PRED_START" else value)

    correlations: set[tuple[str, str]] = set()
    for index, token in enumerate(tokens):
        if token.type != "OPS_COMP" or str(token) != "=":
            continue
        if index == 0 or index == len(tokens) - 1:
            continue

        left, right = tokens[index - 1], tokens[index + 1]
        if left.type == "NAME" and right.type == "PYXFORM_REF":
            question_name = str(right)[2:-1]
            if not question_name.startswith("last-saved#"):
                correlations.add((str(left), question_name))
        elif left.type == "PYXFORM_REF" and right.type == "NAME":
            question_name = str(left)[2:-1]
            if not question_name.startswith("last-saved#"):
                correlations.add((str(right), question_name))

    for index, token in enumerate(tokens):
        if token.type != "FUNC_CALL" or str(token) not in (
            "selected(",
            "contains(",
        ):
            continue
        if index + 4 >= len(tokens):
            continue
        first, comma, second, close = tokens[index + 1 : index + 5]
        if comma.type != "COMMA" or close.type != "CLOSE_PAREN":
            continue
        if first.type == "PYXFORM_REF" and second.type == "NAME":
            question_name = str(first)[2:-1]
            if not question_name.startswith("last-saved#"):
                correlations.add((str(second), question_name))
        elif first.type == "NAME" and second.type == "PYXFORM_REF":
            question_name = str(second)[2:-1]
            if not question_name.startswith("last-saved#"):
                correlations.add((str(first), question_name))

    # This covers common function forms such as selected(${parent}, filter_column).
    if not correlations and len(columns) == 1 and len(questions) == 1:
        correlations.add((next(iter(columns)), next(iter(questions))))

    return columns, questions, correlations


def _survey_row_owner(
    row_type: str,
    name: str,
    row_number: int,
    row_source_map: Mapping[Any, Any] | None,
) -> dict[str, Any]:
    source = _source_for_survey_row(row_source_map, row_number)
    if not source:
        return _generated_survey_owner(row_type, name)
    return {
        key: source[key]
        for key in ("model", "id", "name")
        if source.get(key) is not None
    }


def _expression_question_references(expression: str) -> set[str]:
    references: set[str] = set()
    for token in parse_expression(expression):
        if token.type != "PYXFORM_REF":
            continue
        name = str(token)[2:-1]
        if not name.startswith("last-saved#"):
            references.add(name)
    return references


def _validate_reference_case(
    survey_headers: Mapping[str, int],
    survey_rows: Sequence[tuple[Any, ...]],
    row_source_map: Mapping[Any, Any] | None,
) -> list[ValidationIssue]:
    emitted_names = {
        _row_value(row, survey_headers, "name")
        for row in survey_rows
        if _row_value(row, survey_headers, "name")
    }
    emitted_names_by_casefold: dict[str, set[str]] = defaultdict(set)
    for name in emitted_names:
        emitted_names_by_casefold[name.casefold()].add(name)

    issues: list[ValidationIssue] = []
    expression_columns = [
        column
        for column in survey_headers
        if column not in _REFERENCE_CASE_IGNORED_COLUMNS
    ]
    for row_number, row in enumerate(survey_rows, start=2):
        declaration = _row_value(row, survey_headers, "type")
        owner_name = _row_value(row, survey_headers, "name")
        owner = _survey_row_owner(declaration, owner_name, row_number, row_source_map)

        for column in expression_columns:
            expression = _row_value(row, survey_headers, column)
            if "${" not in expression:
                continue
            for referenced_name in sorted(_expression_question_references(expression)):
                if referenced_name in emitted_names:
                    continue
                available = sorted(
                    emitted_names_by_casefold.get(referenced_name.casefold(), ())
                )
                if not available:
                    # Fully missing references are handled by selected-scope validation.
                    continue
                rendered_available = ", ".join(f"'{name}'" for name in available)
                owner_label = owner.get("name") or owner_name
                issues.append(
                    _issue(
                        "CODEBOOK_REFERENCE_CASE_MISMATCH",
                        "composition",
                        f"{owner['model']} '{owner_label}' field '{column}' references '{referenced_name}', but references are case-sensitive; available exact name: {rendered_available}.",
                        owner=owner,
                        field=column,
                        sheet="survey",
                        column=column,
                        row=row_number,
                    )
                )

    return issues


def _select_list_name(declaration: str) -> tuple[str, str] | None:
    question_type, _, list_name = declaration.partition(" ")
    if question_type not in _INTERNAL_SELECT_TYPES + _EXTERNAL_SELECT_TYPES:
        return None
    return question_type, list_name.strip()


def _validate_choice_filters(
    survey_headers: Mapping[str, int],
    survey_rows: Sequence[tuple[Any, ...]],
    choices_headers: Mapping[str, int],
    choice_values_by_list: Mapping[str, Mapping[str, set[str]]],
    row_source_map: Mapping[Any, Any] | None,
) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    questions: dict[str, dict[str, Any]] = {}
    questions_by_casefold: dict[str, set[str]] = defaultdict(set)

    for row_number, row in enumerate(survey_rows, start=2):
        declaration = _row_value(row, survey_headers, "type")
        name = _row_value(row, survey_headers, "name")
        normalized_type = _normalized_survey_row_type(declaration)
        if (
            not name
            or normalized_type in _END_SURVEY_ROW_TYPES
            or normalized_type in ("begin_group", "begin_repeat")
            or normalized_type in _SURVEY_METADATA_TYPES
        ):
            continue
        questions[name] = {
            "declaration": declaration,
            "row": row_number,
        }
        questions_by_casefold[name.casefold()].add(name)

    columns_by_casefold: dict[str, set[str]] = defaultdict(set)
    for column in choices_headers:
        columns_by_casefold[column.casefold()].add(column)

    for row_number, row in enumerate(survey_rows, start=2):
        expression = _row_value(row, survey_headers, "choice_filter")
        if not expression:
            continue

        declaration = _row_value(row, survey_headers, "type")
        select = _select_list_name(declaration)
        if not select or select[0] in _EXTERNAL_SELECT_TYPES:
            continue

        _, list_name = select
        question_name = _row_value(row, survey_headers, "name")
        owner = _survey_row_owner(
            declaration, question_name, row_number, row_source_map
        )
        source = _source_for_survey_row(row_source_map, row_number) or {}
        columns, referenced_questions, correlations = _choice_filter_references(
            expression
        )

        valid_columns: set[str] = set()
        for column in sorted(columns):
            if column in choices_headers:
                valid_columns.add(column)
                if not choice_values_by_list.get(list_name, {}).get(column):
                    issues.append(
                        _issue(
                            "CODEBOOK_CHOICE_FILTER_COLUMN_VALUES_MISSING",
                            "composition",
                            f"Question '{question_name}' filters choice list '{list_name}' by column '{column}', but that column has no emitted values for the list.",
                            owner=owner,
                            field="choice_filter",
                            sheet="survey",
                            column="choice_filter",
                            row=row_number,
                        )
                    )
                continue

            available = sorted(columns_by_casefold.get(column.casefold(), ()))
            if available:
                code = "CODEBOOK_CHOICE_FILTER_COLUMN_CASE_MISMATCH"
                detail = f"; available column: '{available[0]}'"
            else:
                code = "CODEBOOK_CHOICE_FILTER_COLUMN_NOT_EMITTED"
                detail = ""
            issues.append(
                _issue(
                    code,
                    "composition",
                    f"Question '{question_name}' choice filter references column '{column}', but that exact column is not emitted for choice list '{list_name}'{detail}.",
                    owner=owner,
                    field="choice_filter",
                    sheet="survey",
                    column="choice_filter",
                    row=row_number,
                )
            )

        valid_questions: set[str] = set()
        for referenced_name in sorted(referenced_questions):
            if referenced_name in questions:
                valid_questions.add(referenced_name)
                continue

            available = sorted(
                questions_by_casefold.get(referenced_name.casefold(), ())
            )
            if available:
                code = "CODEBOOK_CHOICE_FILTER_QUESTION_CASE_MISMATCH"
                detail = f"; available question: '{available[0]}'"
            else:
                code = "CODEBOOK_CHOICE_FILTER_QUESTION_NOT_EMITTED"
                detail = ""
            issues.append(
                _issue(
                    code,
                    "composition",
                    f"Question '{question_name}' choice filter references question '{referenced_name}', but that exact question is not available in the generated survey{detail}.",
                    owner=owner,
                    field="choice_filter",
                    sheet="survey",
                    column="choice_filter",
                    row=row_number,
                )
            )

        if columns and referenced_questions and not correlations:
            issues.append(
                _issue(
                    "CODEBOOK_CHOICE_FILTER_UNCORRELATED",
                    "composition",
                    f"Question '{question_name}' choice filter does not clearly correlate its emitted choice columns with its referenced questions.",
                    owner=owner,
                    field="choice_filter",
                    sheet="survey",
                    column="choice_filter",
                    row=row_number,
                )
            )

        configured_filter_list = source.get("choice_filter_list")
        if referenced_questions and "choice_filter_list" in source:
            if not configured_filter_list:
                issues.append(
                    _issue(
                        "CODEBOOK_CHOICE_FILTER_SOURCE_LIST_MISSING",
                        "composition",
                        f"Question '{question_name}' has a choice filter but choice list '{list_name}' does not define the source choice list used by that filter.",
                        owner=owner,
                        field="choice_filter",
                        sheet="survey",
                        column="choice_filter",
                        row=row_number,
                    )
                )

        for filter_column, referenced_name in sorted(correlations):
            if (
                filter_column not in valid_columns
                or referenced_name not in valid_questions
            ):
                continue
            referenced_select = _select_list_name(
                questions[referenced_name]["declaration"]
            )
            if not referenced_select or referenced_select[0] in _EXTERNAL_SELECT_TYPES:
                continue

            referenced_list = referenced_select[1]
            if (
                configured_filter_list
                and referenced_list != configured_filter_list.get("name")
            ):
                issues.append(
                    _issue(
                        "CODEBOOK_CHOICE_FILTER_SOURCE_LIST_INCOMPATIBLE",
                        "composition",
                        f"Question '{question_name}' filters with question '{referenced_name}', which emits choice list '{referenced_list}', but choice list '{list_name}' requires source list '{configured_filter_list.get('name', '')}'.",
                        owner=owner,
                        field="choice_filter",
                        sheet="survey",
                        column="choice_filter",
                        row=row_number,
                    )
                )
                continue
            available_values = choice_values_by_list.get(referenced_list, {}).get(
                "name", set()
            )
            filter_values = choice_values_by_list.get(list_name, {}).get(
                filter_column, set()
            )
            unavailable_values = sorted(filter_values - available_values)
            if unavailable_values:
                rendered_values = ", ".join(
                    f"'{value}'" for value in unavailable_values
                )
                issues.append(
                    _issue(
                        "CODEBOOK_CHOICE_FILTER_VALUE_NOT_EMITTED",
                        "composition",
                        f"Question '{question_name}' choice filter column '{filter_column}' contains {rendered_values}, which are not emitted by question '{referenced_name}' choice list '{referenced_list}'.",
                        owner=owner,
                        field="choice_filter",
                        sheet="survey",
                        column="choice_filter",
                        row=row_number,
                    )
                )

    return issues


def _is_valid_choice_list_name(value: str) -> bool:
    return bool(value) and ":" not in value and is_xml_tag(value)


def _normalized_survey_row_type(value: str) -> str:
    return value.lower().replace(" ", "_")


def _generated_survey_owner(row_type: str, name: str = "") -> dict[str, Any]:
    normalized_type = _normalized_survey_row_type(row_type)
    if normalized_type == "begin_group":
        model = "group"
    elif normalized_type == "begin_repeat":
        model = "repeat"
    elif normalized_type in _SURVEY_METADATA_TYPES:
        model = "metadata"
    else:
        model = "question"

    owner = {"model": model}
    if name:
        owner["name"] = name
    if row_type:
        owner["type"] = row_type
    return owner


def _validate_generated_survey_names(
    survey_headers: Mapping[str, int], survey_rows: Sequence[tuple[Any, ...]]
) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    first_name_rows: dict[str, tuple[int, Mapping[str, Any]]] = {}

    for row_number, row in enumerate(survey_rows, start=2):
        row_type = _row_value(row, survey_headers, "type")
        name = _row_value(row, survey_headers, "name")
        if not row_type and not name:
            continue

        normalized_type = _normalized_survey_row_type(row_type)
        if normalized_type in _END_SURVEY_ROW_TYPES:
            continue

        owner = _generated_survey_owner(row_type, name)
        if not name:
            issues.append(
                _issue(
                    "CODEBOOK_GENERATED_NAME_MISSING",
                    "composition",
                    f"Survey row {row_number} of type '{row_type}' requires a generated name.",
                    owner=owner,
                    field="name",
                    sheet="survey",
                    column="name",
                    row=row_number,
                )
            )
            continue

        if name in _RESERVED_SURVEY_NAMES:
            issues.append(
                _issue(
                    "CODEBOOK_GENERATED_NAME_INVALID",
                    "composition",
                    f"Generated {owner['model']} name '{name}' is reserved and cannot be emitted on the survey sheet.",
                    owner=owner,
                    field="name",
                    sheet="survey",
                    column="name",
                    row=row_number,
                )
            )
        elif ":" in name or not is_xml_tag(name):
            issues.append(
                _issue(
                    "CODEBOOK_GENERATED_NAME_INVALID",
                    "composition",
                    f"Generated {owner['model']} name '{name}' is invalid. Names must begin with a letter or underscore and contain only letters, digits, underscores, hyphens, or periods.",
                    owner=owner,
                    field="name",
                    sheet="survey",
                    column="name",
                    row=row_number,
                )
            )

        first_name = first_name_rows.setdefault(name, (row_number, owner))
        first_row, first_owner = first_name
        if first_row != row_number:
            duplicate_owner = dict(owner)
            duplicate_owner["first_model"] = first_owner["model"]
            duplicate_owner["first_row"] = first_row
            issues.append(
                _issue(
                    "CODEBOOK_GENERATED_NAME_DUPLICATE",
                    "composition",
                    f"Generated {owner['model']} name '{name}' duplicates a {first_owner['model']} first emitted at survey row {first_row}.",
                    owner=duplicate_owner,
                    field="name",
                    sheet="survey",
                    column="name",
                    row=row_number,
                )
            )

    return issues


def _base_question_type(declaration: str) -> str:
    return declaration.partition(" ")[0]


def _validate_generated_survey_types(
    survey_headers: Mapping[str, int], survey_rows: Sequence[tuple[Any, ...]]
) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    for row_number, row in enumerate(survey_rows, start=2):
        declaration = _row_value(row, survey_headers, "type")
        name = _row_value(row, survey_headers, "name")
        normalized_type = _normalized_survey_row_type(declaration)
        if not declaration and name:
            issues.append(
                _issue(
                    "CODEBOOK_GENERATED_TYPE_UNSUPPORTED",
                    "composition",
                    f"Generated question '{name}' has no type.",
                    owner=_generated_survey_owner(declaration, name),
                    field="type",
                    sheet="survey",
                    column="type",
                    row=row_number,
                )
            )
            continue
        if (
            not declaration
            or normalized_type in _END_SURVEY_ROW_TYPES
            or normalized_type in ("begin_group", "begin_repeat")
            or normalized_type in _SURVEY_METADATA_TYPES
        ):
            continue

        question_type = _base_question_type(declaration)
        if question_type not in _SUPPORTED_GENERATED_TYPES:
            issues.append(
                _issue(
                    "CODEBOOK_GENERATED_TYPE_UNSUPPORTED",
                    "composition",
                    f"Generated question '{name}' uses unsupported type '{question_type}'.",
                    owner=_generated_survey_owner(declaration, name),
                    field="type",
                    sheet="survey",
                    column="type",
                    row=row_number,
                )
            )
    return issues


def _split_emitted_names(value: str) -> set[str]:
    return {name.strip() for name in value.split(",") if name.strip()}


def _validate_codebook_export_definitions(
    workbook: Any,
    survey_headers: Mapping[str, int],
    survey_rows: Sequence[tuple[Any, ...]],
    emitted_lists: set[str],
) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    suffixes: dict[str, dict[str, Any]] = {}
    recall_periods: dict[str, dict[str, Any]] = {}

    suffix_headers: Mapping[str, int] = {}
    suffix_rows: Sequence[tuple[Any, ...]] = ()
    if "suffixes" in workbook.sheetnames:
        suffix_headers, suffix_rows = _sheet_rows_by_header(workbook["suffixes"])

    for row_number, row in enumerate(suffix_rows, start=2):
        name = _row_value(row, suffix_headers, "name")
        suffix_type = _row_value(row, suffix_headers, "type")
        choice_list = _row_value(row, suffix_headers, "choicelist")
        nested_suffixes = _split_emitted_names(
            _row_value(row, suffix_headers, "suffix")
        )
        if not any((name, suffix_type, choice_list, nested_suffixes)):
            continue

        owner = {"model": "suffix", "name": name} if name else {"model": "suffix"}
        if not name:
            issues.append(
                _issue(
                    "CODEBOOK_SUFFIX_NAME_MISSING",
                    "composition",
                    "An emitted suffix definition has no name.",
                    owner=owner,
                    field="name",
                    sheet="suffixes",
                    column="name",
                    row=row_number,
                )
            )
        else:
            if ":" in name or not is_xml_tag(name):
                issues.append(
                    _issue(
                        "CODEBOOK_SUFFIX_NAME_INVALID",
                        "composition",
                        f"Suffix name '{name}' is invalid. Names must begin with a letter or underscore and contain only letters, digits, underscores, hyphens, or periods.",
                        owner=owner,
                        field="name",
                        sheet="suffixes",
                        column="name",
                        row=row_number,
                    )
                )
            if name in suffixes:
                issues.append(
                    _issue(
                        "CODEBOOK_SUFFIX_NAME_DUPLICATE",
                        "composition",
                        f"Suffix '{name}' is duplicated; it was first emitted at row {suffixes[name]['row']}.",
                        owner={
                            "model": "suffix",
                            "name": name,
                            "first_row": suffixes[name]["row"],
                        },
                        field="name",
                        sheet="suffixes",
                        column="name",
                        row=row_number,
                    )
                )
            else:
                suffixes[name] = {
                    "type": suffix_type,
                    "choice_list": choice_list,
                    "nested_suffixes": nested_suffixes,
                    "row": row_number,
                }

        if suffix_type not in _SUPPORTED_GENERATED_TYPES:
            issues.append(
                _issue(
                    "CODEBOOK_SUFFIX_TYPE_UNSUPPORTED",
                    "composition",
                    f"Suffix '{name}' uses unsupported type '{suffix_type}'.",
                    owner=owner,
                    field="type",
                    sheet="suffixes",
                    column="type",
                    row=row_number,
                )
            )
        elif suffix_type in _INTERNAL_SELECT_TYPES:
            if not choice_list:
                issues.append(
                    _issue(
                        "CODEBOOK_SUFFIX_CHOICE_LIST_MISSING",
                        "composition",
                        f"Suffix '{name}' uses {suffix_type} but has no choice list.",
                        owner=owner,
                        field="choices",
                        sheet="suffixes",
                        column="choicelist",
                        row=row_number,
                    )
                )
            elif choice_list not in emitted_lists:
                issues.append(
                    _issue(
                        "CODEBOOK_SUFFIX_CHOICE_LIST_NOT_EMITTED",
                        "composition",
                        f"Suffix '{name}' references choice list '{choice_list}', but that list has no emitted choices.",
                        owner=owner,
                        field="choices",
                        sheet="suffixes",
                        column="choicelist",
                        row=row_number,
                    )
                )
        elif suffix_type not in _EXTERNAL_SELECT_TYPES and choice_list:
            issues.append(
                _issue(
                    "CODEBOOK_SUFFIX_CHOICE_LIST_INCOMPATIBLE",
                    "composition",
                    f"Suffix '{name}' has type '{suffix_type}', which cannot use choice list '{choice_list}'.",
                    owner=owner,
                    field="choices",
                    sheet="suffixes",
                    column="choicelist",
                    row=row_number,
                )
            )

    recall_headers: Mapping[str, int] = {}
    recall_rows: Sequence[tuple[Any, ...]] = ()
    if "recall_periods" in workbook.sheetnames:
        recall_headers, recall_rows = _sheet_rows_by_header(workbook["recall_periods"])

    for row_number, row in enumerate(recall_rows, start=2):
        name = _row_value(row, recall_headers, "name")
        description = _row_value(row, recall_headers, "description")
        if not name and not description:
            continue

        owner = (
            {"model": "recall_period", "name": name}
            if name
            else {"model": "recall_period"}
        )
        if not name:
            issues.append(
                _issue(
                    "CODEBOOK_RECALL_PERIOD_NAME_MISSING",
                    "composition",
                    "An emitted recall-period definition has no name.",
                    owner=owner,
                    field="name",
                    sheet="recall_periods",
                    column="name",
                    row=row_number,
                )
            )
            continue
        if ":" in name or not is_xml_tag(name):
            issues.append(
                _issue(
                    "CODEBOOK_RECALL_PERIOD_NAME_INVALID",
                    "composition",
                    f"Recall-period name '{name}' is invalid. Names must begin with a letter or underscore and contain only letters, digits, underscores, hyphens, or periods.",
                    owner=owner,
                    field="name",
                    sheet="recall_periods",
                    column="name",
                    row=row_number,
                )
            )
        if name in recall_periods:
            issues.append(
                _issue(
                    "CODEBOOK_RECALL_PERIOD_NAME_DUPLICATE",
                    "composition",
                    f"Recall period '{name}' is duplicated; it was first emitted at row {recall_periods[name]['row']}.",
                    owner={
                        "model": "recall_period",
                        "name": name,
                        "first_row": recall_periods[name]["row"],
                    },
                    field="name",
                    sheet="recall_periods",
                    column="name",
                    row=row_number,
                )
            )
        else:
            recall_periods[name] = {"row": row_number}

    for row_number, row in enumerate(survey_rows, start=2):
        question_name = _row_value(row, survey_headers, "name")
        suffix_1_name = _row_value(row, survey_headers, "suffix1")
        suffix_2_name = _row_value(row, survey_headers, "suffix2")
        recall_period_name = _row_value(row, survey_headers, "recall_period")
        if not any((suffix_1_name, suffix_2_name, recall_period_name)):
            continue

        owner = {"model": "question", "name": question_name}
        suffix_1 = suffixes.get(suffix_1_name)
        suffix_2 = suffixes.get(suffix_2_name)
        if suffix_1_name and not suffix_1:
            issues.append(
                _issue(
                    "CODEBOOK_SUFFIX_NOT_EMITTED",
                    "composition",
                    f"Question '{question_name}' references suffix '{suffix_1_name}', but that suffix was not emitted.",
                    owner=owner,
                    field="suffix1",
                    sheet="survey",
                    column="suffix1",
                    row=row_number,
                )
            )
        if suffix_2_name and not suffix_2:
            issues.append(
                _issue(
                    "CODEBOOK_SUFFIX_NOT_EMITTED",
                    "composition",
                    f"Question '{question_name}' references suffix '{suffix_2_name}', but that suffix was not emitted.",
                    owner=owner,
                    field="suffix2",
                    sheet="survey",
                    column="suffix2",
                    row=row_number,
                )
            )
        if suffix_2_name and not suffix_1_name:
            issues.append(
                _issue(
                    "CODEBOOK_NESTED_SUFFIX_INVALID",
                    "composition",
                    f"Question '{question_name}' uses suffix 2 '{suffix_2_name}' without suffix 1.",
                    owner=owner,
                    field="suffix2",
                    sheet="survey",
                    column="suffix2",
                    row=row_number,
                )
            )
        elif suffix_1 and suffix_2 and suffix_2_name not in suffix_1["nested_suffixes"]:
            issues.append(
                _issue(
                    "CODEBOOK_NESTED_SUFFIX_INVALID",
                    "composition",
                    f"Question '{question_name}' uses suffix 2 '{suffix_2_name}', which is not nested under suffix 1 '{suffix_1_name}'.",
                    owner=owner,
                    field="suffix2",
                    sheet="survey",
                    column="suffix2",
                    row=row_number,
                )
            )

        if recall_period_name and recall_period_name not in recall_periods:
            issues.append(
                _issue(
                    "CODEBOOK_RECALL_PERIOD_NOT_EMITTED",
                    "composition",
                    f"Question '{question_name}' references recall period '{recall_period_name}', but that recall period was not emitted.",
                    owner=owner,
                    field="recall_period",
                    sheet="survey",
                    column="recall_period",
                    row=row_number,
                )
            )

        effective_suffix = suffix_2 or suffix_1
        if not effective_suffix:
            continue
        emitted_type = _row_value(row, survey_headers, "type")
        emitted_choice_list = _row_value(row, survey_headers, "choice_list")
        if emitted_type != effective_suffix["type"]:
            issues.append(
                _issue(
                    "CODEBOOK_SUFFIX_TYPE_INCOMPATIBLE",
                    "composition",
                    f"Question '{question_name}' emits type '{emitted_type}', but its effective suffix requires type '{effective_suffix['type']}'.",
                    owner=owner,
                    field="type",
                    sheet="survey",
                    column="type",
                    row=row_number,
                )
            )
        if (
            effective_suffix["type"] in _INTERNAL_SELECT_TYPES
            and emitted_choice_list != effective_suffix["choice_list"]
        ):
            issues.append(
                _issue(
                    "CODEBOOK_SUFFIX_CHOICE_LIST_INCOMPATIBLE",
                    "composition",
                    f"Question '{question_name}' emits choice list '{emitted_choice_list}', but its effective suffix requires '{effective_suffix['choice_list']}'.",
                    owner=owner,
                    field="choices",
                    sheet="survey",
                    column="choice_list",
                    row=row_number,
                )
            )

    return issues


def _source_for_survey_row(
    row_source_map: Mapping[Any, Any] | None, row_number: int
) -> Mapping[str, Any] | None:
    if not row_source_map:
        return None
    return row_source_map.get(("survey", row_number)) or row_source_map.get(
        f"survey:{row_number}"
    )


def _validate_final_source_context(
    survey_headers: Mapping[str, int],
    survey_rows: Sequence[tuple[Any, ...]],
    row_source_map: Mapping[Any, Any] | None,
) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    for row_number, row in enumerate(survey_rows, start=2):
        source = _source_for_survey_row(row_source_map, row_number)
        if not source or source.get("model") != "SubQuestion":
            continue

        owner = {
            key: source[key]
            for key in ("model", "id", "name")
            if source.get(key) is not None
        }
        suffix_1 = source.get("suffix")
        suffix_2 = source.get("suffix_2")
        recall_period = source.get("recall_period")
        if suffix_2 and not suffix_1:
            issues.append(
                _issue(
                    "CODEBOOK_NESTED_SUFFIX_INVALID",
                    "composition",
                    f"Question '{source.get('name', '')}' uses suffix 2 '{suffix_2.get('name', '')}' without suffix 1.",
                    owner=owner,
                    field="suffix_2",
                    sheet="survey",
                    column="name",
                    row=row_number,
                )
            )
        elif (
            suffix_1
            and suffix_2
            and suffix_2.get("name") not in suffix_1.get("nested_suffixes", ())
        ):
            issues.append(
                _issue(
                    "CODEBOOK_NESTED_SUFFIX_INVALID",
                    "composition",
                    f"Question '{source.get('name', '')}' uses suffix 2 '{suffix_2.get('name', '')}', which is not nested under suffix 1 '{suffix_1.get('name', '')}'.",
                    owner=owner,
                    field="suffix_2",
                    sheet="survey",
                    column="name",
                    row=row_number,
                )
            )

        for reference, missing_code, invalid_code, field_name, label in (
            (
                suffix_1,
                "CODEBOOK_SUFFIX_NAME_MISSING",
                "CODEBOOK_SUFFIX_NAME_INVALID",
                "suffix",
                "Suffix 1",
            ),
            (
                suffix_2,
                "CODEBOOK_SUFFIX_NAME_MISSING",
                "CODEBOOK_SUFFIX_NAME_INVALID",
                "suffix_2",
                "Suffix 2",
            ),
            (
                recall_period,
                "CODEBOOK_RECALL_PERIOD_NAME_MISSING",
                "CODEBOOK_RECALL_PERIOD_NAME_INVALID",
                "recall_period",
                "Recall period",
            ),
        ):
            if reference is not None and not reference.get("name"):
                issues.append(
                    _issue(
                        missing_code,
                        "composition",
                        f"{label} for question '{source.get('name', '')}' has no name.",
                        owner=owner,
                        field=field_name,
                        sheet="survey",
                        column="name",
                        row=row_number,
                    )
                )
            elif reference is not None and (
                ":" in reference["name"] or not is_xml_tag(reference["name"])
            ):
                reference_owner = {
                    key: reference[key]
                    for key in ("model", "id", "name")
                    if reference.get(key) is not None
                }
                issues.append(
                    _issue(
                        invalid_code,
                        "composition",
                        f"{label} name '{reference['name']}' is invalid.",
                        owner=reference_owner,
                        field="name",
                        sheet="survey",
                        column="name",
                        row=row_number,
                    )
                )

        expected_type = source.get("effective_type", "")
        emitted_declaration = _row_value(row, survey_headers, "type")
        emitted_type = _base_question_type(emitted_declaration)
        if expected_type and emitted_type != expected_type:
            issues.append(
                _issue(
                    "CODEBOOK_SUFFIX_TYPE_INCOMPATIBLE",
                    "composition",
                    f"Question '{source.get('name', '')}' emits type '{emitted_type}', but its selected suffix context requires '{expected_type}'.",
                    owner=owner,
                    field="type",
                    sheet="survey",
                    column="type",
                    row=row_number,
                )
            )

        expected_choice_list = source.get("effective_choice_list", "")
        declaration = _internal_select_declaration(row, survey_headers)
        if declaration and expected_choice_list:
            _, emitted_choice_list, list_column = declaration
            if emitted_choice_list != expected_choice_list:
                issues.append(
                    _issue(
                        "CODEBOOK_SUFFIX_CHOICE_LIST_INCOMPATIBLE",
                        "composition",
                        f"Question '{source.get('name', '')}' emits choice list '{emitted_choice_list}', but its selected suffix context requires '{expected_choice_list}'.",
                        owner=owner,
                        field="choices",
                        sheet="survey",
                        column=list_column,
                        row=row_number,
                    )
                )

    return issues


def validate_codebook_integrity(
    xlsx_bytes: bytes, row_source_map: Mapping[Any, Any] | None = None
) -> list[ValidationIssue]:
    """Validate internal select declarations and exact emitted choice rows."""

    workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    try:
        if "survey" not in workbook.sheetnames or "choices" not in workbook.sheetnames:
            return []

        survey_headers, survey_rows = _sheet_rows_by_header(workbook["survey"])
        choices_headers, choices_rows = _sheet_rows_by_header(workbook["choices"])
        is_codebook_export = "choice_list" in survey_headers
        issues: list[ValidationIssue] = []
        if not is_codebook_export:
            issues.extend(_validate_generated_survey_names(survey_headers, survey_rows))
            issues.extend(_validate_generated_survey_types(survey_headers, survey_rows))
            issues.extend(
                _validate_final_source_context(
                    survey_headers, survey_rows, row_source_map
                )
            )
            issues.extend(
                _validate_reference_case(survey_headers, survey_rows, row_source_map)
            )
        emitted_list_column = (
            "list_name" if "list_name" in choices_headers else "choice_list"
        )
        english_label_column = _english_label_column(choices_headers)
        internal_select_types_by_list: dict[str, set[str]] = defaultdict(set)
        for row in survey_rows:
            declaration = _internal_select_declaration(row, survey_headers)
            if declaration:
                question_type, list_name, _ = declaration
                if list_name:
                    internal_select_types_by_list[list_name].add(question_type)

        emitted_lists: set[str] = set()
        choice_values_by_list: dict[str, dict[str, set[str]]] = defaultdict(
            lambda: defaultdict(set)
        )
        first_choice_value_rows: dict[tuple[str, str], int] = {}
        checked_list_names: set[str] = set()
        reported_invalid_list_names: set[str] = set()
        for row_number, row in enumerate(choices_rows, start=2):
            list_name = _row_value(row, choices_headers, emitted_list_column)
            choice_name = _row_value(row, choices_headers, "name")
            english_label = (
                _row_value(row, choices_headers, english_label_column)
                if english_label_column
                else ""
            )
            if not any((list_name, choice_name, english_label)):
                continue
            if not list_name:
                issues.append(
                    _issue(
                        "CODEBOOK_CHOICE_LIST_NAME_MISSING",
                        "composition",
                        "An emitted choice row does not specify a choice list name.",
                        owner={"model": "choice", "name": choice_name},
                        field=emitted_list_column,
                        sheet="choices",
                        column=emitted_list_column,
                        row=row_number,
                    )
                )
            else:
                emitted_lists.add(list_name)
                for choice_column in choices_headers:
                    value = _row_value(row, choices_headers, choice_column)
                    if value:
                        choice_values_by_list[list_name][choice_column].add(value)
                if list_name not in checked_list_names:
                    checked_list_names.add(list_name)
                    if not _is_valid_choice_list_name(list_name):
                        reported_invalid_list_names.add(list_name)
                        issues.append(
                            _issue(
                                "CODEBOOK_CHOICE_LIST_NAME_INVALID",
                                "composition",
                                f"Choice list name '{list_name}' is invalid. Names must begin with a letter or underscore and contain only letters, digits, underscores, hyphens, or periods.",
                                owner={"model": "choice_list", "name": list_name},
                                field=emitted_list_column,
                                sheet="choices",
                                column=emitted_list_column,
                                row=row_number,
                            )
                        )

            if not choice_name:
                issues.append(
                    _issue(
                        "CODEBOOK_CHOICE_VALUE_MISSING",
                        "composition",
                        f"Choice list '{list_name}' has an emitted choice with no value.",
                        owner={"model": "choice_list", "name": list_name},
                        field="name",
                        sheet="choices",
                        column="name",
                        row=row_number,
                    )
                )
            elif list_name:
                duplicate_key = (list_name, choice_name)
                first_row = first_choice_value_rows.setdefault(
                    duplicate_key, row_number
                )
                if first_row != row_number:
                    issues.append(
                        _issue(
                            "CODEBOOK_CHOICE_VALUE_DUPLICATE",
                            "composition",
                            f"Choice value '{choice_name}' is duplicated in choice list '{list_name}'; it was first emitted at row {first_row}.",
                            owner={
                                "model": "choice",
                                "name": choice_name,
                                "choice_list": list_name,
                            },
                            field="name",
                            sheet="choices",
                            column="name",
                            row=row_number,
                        )
                    )
                if "select_multiple" in internal_select_types_by_list[
                    list_name
                ] and any(character.isspace() for character in choice_name):
                    issues.append(
                        _issue(
                            "CODEBOOK_CHOICE_VALUE_INVALID",
                            "composition",
                            f"Choice value '{choice_name}' in choice list '{list_name}' contains whitespace, which is not supported by select_multiple questions.",
                            owner={
                                "model": "choice",
                                "name": choice_name,
                                "choice_list": list_name,
                            },
                            field="name",
                            sheet="choices",
                            column="name",
                            row=row_number,
                        )
                    )
            if english_label_column and not english_label:
                issues.append(
                    _issue(
                        "CODEBOOK_CHOICE_LABEL_MISSING",
                        "composition",
                        f"Choice '{choice_name}' in choice list '{list_name}' has no English label.",
                        owner={
                            "model": "choice",
                            "name": choice_name,
                            "choice_list": list_name,
                        },
                        field="label",
                        sheet="choices",
                        column=english_label_column,
                        row=row_number,
                    )
                )

        if is_codebook_export:
            issues.extend(
                _validate_codebook_export_definitions(
                    workbook, survey_headers, survey_rows, emitted_lists
                )
            )
        else:
            issues.extend(
                _validate_choice_filters(
                    survey_headers,
                    survey_rows,
                    choices_headers,
                    choice_values_by_list,
                    row_source_map,
                )
            )

        for row_number, row in enumerate(survey_rows, start=2):
            declaration = _internal_select_declaration(row, survey_headers)
            if not declaration:
                continue

            question_type, list_name, list_column = declaration
            question_name = _row_value(row, survey_headers, "name")
            owner = {"model": "question", "name": question_name}

            if not list_name:
                issues.append(
                    _issue(
                        "CODEBOOK_CHOICE_LIST_MISSING",
                        "composition",
                        f"Question '{question_name}' uses {question_type} but does not specify a choice list.",
                        owner=owner,
                        field="choices",
                        sheet="survey",
                        column=list_column,
                        row=row_number,
                    )
                )
                continue

            if (
                not _is_valid_choice_list_name(list_name)
                and list_name not in reported_invalid_list_names
            ):
                reported_invalid_list_names.add(list_name)
                issues.append(
                    _issue(
                        "CODEBOOK_CHOICE_LIST_NAME_INVALID",
                        "composition",
                        f"Choice list name '{list_name}' is invalid. Names must begin with a letter or underscore and contain only letters, digits, underscores, hyphens, or periods.",
                        owner=owner,
                        field="choices",
                        sheet="survey",
                        column=list_column,
                        row=row_number,
                    )
                )

            if list_name not in emitted_lists:
                issues.append(
                    _issue(
                        "CODEBOOK_CHOICE_LIST_NOT_EMITTED",
                        "composition",
                        f"Question '{question_name}' references choice list '{list_name}', but that list has no emitted choices.",
                        owner=owner,
                        field="choices",
                        sheet="survey",
                        column=list_column,
                        row=row_number,
                    )
                )
        return issues
    finally:
        workbook.close()


def build_generated_artifact(xlsx_form: Any) -> GeneratedSurveyArtifact:
    """Generate a workbook once and materialize its external files once."""

    try:
        xlsx_bytes = xlsx_form.generate()
    except Exception as exc:
        raise ArtifactInfrastructureError(f"Unable to generate XLSX: {exc}") from exc

    external_files = materialize_external_files(xlsx_form.external_files)
    return GeneratedSurveyArtifact(
        xlsx_bytes=xlsx_bytes,
        external_files=external_files,
        row_source_map=getattr(xlsx_form, "row_source_map", None),
        form_name=str(xlsx_form.id_name or "survey"),
    )


def validate_generated_artifact(
    artifact: GeneratedSurveyArtifact,
    *,
    converter_cls: Any | None = None,
) -> ValidationResult:
    """Convert and validate one exact artifact without generating it again."""

    try:
        composition_errors = validate_codebook_integrity(
            artifact.xlsx_bytes, artifact.row_source_map
        )
    except Exception as exc:
        raise ValidatorInfrastructureError(
            f"Unable to inspect the generated XLSX codebook: {exc}"
        ) from exc

    if composition_errors:
        return ValidationResult(
            valid=False,
            artifact_hash=artifact.artifact_hash,
            errors=composition_errors,
            artifact=artifact,
        )

    if converter_cls is None:
        from .xml_conversion import XMLConversion

        converter_cls = XMLConversion

    try:
        conversion = converter_cls(io.BytesIO(artifact.xlsx_bytes))
        xml = conversion.run()
    except Exception as exc:
        raise ValidatorInfrastructureError(
            f"pyxform conversion failed internally: {exc}"
        ) from exc

    errors = _normalise_messages(conversion.errors, warning=False)
    warnings = _normalise_messages(conversion.warnings, warning=True)
    if not errors:
        errors.extend(validate_xml_compatibility(xml, artifact.external_files))

    validated_artifact = replace(artifact, xml=xml)
    return ValidationResult(
        valid=not errors,
        artifact_hash=validated_artifact.artifact_hash or artifact.artifact_hash,
        errors=errors,
        warnings=warnings,
        artifact=validated_artifact,
    )


def failed_validation_result(
    issue: ValidationIssue, *, artifact_hash: str = _EMPTY_ARTIFACT_HASH
) -> ValidationResult:
    return ValidationResult(valid=False, artifact_hash=artifact_hash, errors=(issue,))


__all__ = [
    "ArtifactInfrastructureError",
    "ArtifactInputError",
    "COMPATIBILITY_VERSION",
    "GeneratedSurveyArtifact",
    "PYXFORM_VERSION",
    "ValidatorInfrastructureError",
    "ValidationIssue",
    "ValidationResult",
    "build_generated_artifact",
    "compute_artifact_hash",
    "failed_validation_result",
    "materialize_external_files",
    "validate_codebook_integrity",
    "validate_generated_artifact",
    "validate_xml_compatibility",
]

import enum
import string

from django.conf import settings
from django.db.models import Prefetch
from django.db.models.functions import Coalesce
from modules.models import Indicator, Submodule, SubmoduleRequiredGroup
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.writer.excel import save_virtual_workbook
from questions.models import (
    BaseQuestion,
    Choice,
    ChoiceGroup,
    ChoiceTranslation,
    RecallPeriod,
    RepeatSection,
    RepeatSectionTranslation,
    RootQuestion,
    RootQuestionTranslation,
    SubQuestion,
    SubQuestionTranslation,
    Suffix,
)
from questions.serializers import (
    RepeatSectionExportSerializer,
    RootQuestionExportSerializer,
    SubQuestionExportSerializer,
)


class QuestionsExport:
    STATIC_COLUMNS = (
        "module_name",
        "module_label",
        "submodule_name",
        "submodule_label",
        "type",
        "choice_list",
        "name",
        "label",
        "hint",
        "suffix1",
        "suffix2",
        "recall_period",
        "relevant",
        "constraint",
        "constraint_message",
        "calculation",
        "description",
        "repeat",
        "repeat_count",
        "appearance",
        "parameters",
        "default",
        "disabled",
        "required",
        "read_only",
    )
    LANGUAGE_COLUMNS = ("label", "hint")

    COLUMN_COMMENT = {
        "module_name": "mandatory",
        "module_label": "mandatory",
        "submodule_name": "mandatory",
        "submodule_label": "mandatory",
        "type": "mandatory",
        "choice_list": "mandatory if type = select one or type = select multiple",
        "name": "mandatory",
        "label": "mandatory",
        "hint": "optional",
        "suffix": "optional \n comma separated list of nested suffixes e.g _oth,_YN",
        "suffix1": "optional",
        "suffix2": "optional",
        "recall_period": "optional",
        "relevant": "optional",
        "constraint": "optional",
        "constraint_message": "optional",
        "calculation": "mandatory if type = calculate",
        "description": "optional",
        "repeat": "optional",
        "repeat_count": "required if type = repeat",
        "appearance": "optional",
        "parameters": "optional",
        "name": "mandatory",
        "choicelist": "mandatory if type = select one or type = select multiple",
        "indicator_name": "mandatory",
        "indicator_label": "mandatory",
        "question_name": "mandatory",
    }

    ENGLISH_LANGUAGES = {
        "en": "English",
        "fr": "French",
        "es": "Spanish",
        "ar": "Arabic",
        "zh-cn": "Chinese",
        "ru": "Russian",
        "pt": "Portuguese",
    }

    def __init__(self, languages=[]):
        self.IS_TEMPLATE = False
        self.wb = Workbook()
        self.questions_sheet = self.wb.active
        self.questions_sheet.title = "survey"
        self.choices_sheet = self.wb.create_sheet(title="choices")
        self.suffixes_sheet = self.wb.create_sheet(title="suffixes")
        self.recall_period_sheet = self.wb.create_sheet(title="recall_periods")
        self.required_groups_sheet = self.wb.create_sheet(title="required_groups")
        self.indicators_sheet = self.wb.create_sheet(title="indicators")
        self.current_row_index = 1
        self.repeat_section_ids = set()
        self.choices_ids = set()
        self.suffixes_ids = set()
        self.recall_period_ids = set()
        self.indicator_ids = set()
        self.base_question_ids = set()
        self.repeat_section_added = set()
        self.required_group_ids = set()
        self.choice_groups = []

        self.languages_dict = dict(settings.LANGUAGES)
        self.language_values = set()
        self.languages = languages

        self.bold_font = Font(bold=True)
        self.columns = []

    def _add_comment_to_column(self, column, cell):
        if self.IS_TEMPLATE:
            column_name = column.name.split("::")[0]
            if column_name in self.COLUMN_COMMENT:
                cell.comment = Comment(self.COLUMN_COMMENT[column_name], "System")

    def increment_row_index(self):
        self.current_row_index += 1

    @staticmethod
    def get_language_column(name, language):
        value, display = language
        return f"{name}::{display} ({value})"

    def get_choices(self):
        choices_translations_prefetch = Prefetch(
            "translations",
            queryset=ChoiceTranslation.objects.all(),
        )
        choices_prefetch = Prefetch(
            "choices",
            queryset=Choice.objects.filter(is_active=True)
            .prefetch_related(choices_translations_prefetch)
            .order_by("order", "id"),
        )
        return ChoiceGroup.objects.filter(id__in=self.choice_groups).prefetch_related(
            choices_prefetch
        )

    @staticmethod
    def get_suffixes(ids):
        return Suffix.objects.filter(id__in=ids)

    @staticmethod
    def get_recall_periods(ids):
        return RecallPeriod.objects.filter(id__in=ids)

    @staticmethod
    def get_required_groups(ids):
        return SubmoduleRequiredGroup.objects.filter(id__in=ids)

    @staticmethod
    def get_indicators(ids):
        return Indicator.objects.filter(id__in=ids)

    def _generate_columns(self):
        columns = []

        for column in self.STATIC_COLUMNS:
            if self.languages and column in self.LANGUAGE_COLUMNS:
                for language in self.languages:
                    columns.append(self.get_language_column(column, language))
            else:
                columns.append(column)
        self.columns = enum.Enum("Columns", columns)

    def add_columns(self):
        for column in self.columns:
            cell = self.questions_sheet.cell(
                row=1, column=column.value, value=column.name
            )
            cell.font = self.bold_font
            self._add_comment_to_column(column, cell)

    def get_column_value(self, name, language=None):
        key = self.get_language_column(name, language) if language else name
        try:
            return self.columns[key].value
        except KeyError:
            # TODO: add logging
            return

    def fill_cell(self, sheet, column_name, value, bold=False, language: tuple = None):
        column = self.get_column_value(column_name, language=language)

        if not column:
            # TODO: add logging
            return

        cell = sheet.cell(row=self.current_row_index, column=column, value=value)

        if bold:
            cell.font = self.bold_font
        return cell

    def _populate_rows(self, data):
        for key, value in data.items():
            if key == "translations":
                for translation in value:
                    language = (
                        translation["language"],
                        translation["language_display"],
                    )
                    self.fill_cell(
                        self.questions_sheet,
                        "label",
                        translation["label"],
                        language=language,
                    )
                    if translation.get("hint"):
                        self.fill_cell(
                            self.questions_sheet,
                            "hint",
                            translation["hint"],
                            language=language,
                        )
            else:
                self.fill_cell(self.questions_sheet, key, value)

    def add_repeat_section(self, repeat_section):
        if repeat_section in self.repeat_section_added:
            return

        data = RepeatSectionExportSerializer(instance=repeat_section).data
        submodule_names = data.get("submodule_name", "").split(", ")
        submodule_labels = data.get("submodule_label", "").split(", ")
        module_names = data.get("module_name", "").split(", ")
        module_labels = data.get("module_label", "").split(", ")

        for i in range(len(submodule_names)):
            data["submodule_name"] = submodule_names[i]
            data["submodule_label"] = submodule_labels[i]
            data["module_name"] = module_names[i]
            data["module_label"] = module_labels[i]
            self.increment_row_index()
            self._populate_rows(data)

        self.repeat_section_added.add(repeat_section)

    def add_question_row(self, question, skip_required_groups=False):
        if isinstance(question, BaseQuestion):
            base_question = question
            question = question.instance
        else:
            base_question = question.base_question

        if isinstance(question, RootQuestion):
            data = RootQuestionExportSerializer(instance=question).data
        elif isinstance(question, SubQuestion):
            data = SubQuestionExportSerializer(instance=question).data
            if not skip_required_groups:
                submodule_required_group_ids = (
                    question.root_question.submodule.values_list(
                        "required_groups", flat=True
                    )
                )
                relevant_required_group_ids = (
                    SubmoduleRequiredGroup.objects.filter(
                        id__in=submodule_required_group_ids
                    )
                    .filter(
                        required_suffix=question.suffix,
                        required_nested_suffix=question.suffix_2,
                        required_recall_period=question.recall_period,
                    )
                    .values_list("id", flat=True)
                )
                self.required_group_ids.update(relevant_required_group_ids)

        elif isinstance(question, RepeatSection):
            self.add_repeat_section(question)
            return
        else:
            raise ValueError("Provided question could not be serialized.")

        indicators = Indicator.objects.filter(questions=base_question).values_list(
            "id", flat=True
        )
        self.indicator_ids.update(indicators)

        repeat_sections = base_question.repeat_sections.values_list("name", flat=True)

        submodule_names = data.get("submodule_name", "").split(", ")
        submodule_labels = data.get("submodule_label", "").split(", ")
        module_names = data.get("module_name", "").split(", ")
        module_labels = data.get("module_label", "").split(", ")

        for i in range(len(submodule_names)):
            data["submodule_name"] = submodule_names[i]
            data["submodule_label"] = submodule_labels[i]
            data["module_name"] = module_names[i]
            data["module_label"] = module_labels[i]
            self.increment_row_index()
            self._populate_rows(data)

        if repeat_sections:
            self.fill_cell(self.questions_sheet, "repeat", ",".join(repeat_sections))

        if isinstance(question, SubQuestion):
            if suffix := question.suffix:
                self.suffixes_ids.add(suffix.id)
            if suffix_2 := question.suffix_2:
                self.suffixes_ids.add(suffix_2.id)
            if recall_period := question.recall_period:
                self.recall_period_ids.add(recall_period.id)

        self.base_question_ids.add(base_question.id)

    def add_choices(self):
        columns = ["choice_list", "name"]

        for language in self.languages:
            value, display = language
            columns.append(f"label::{display} ({value})")

        row = 1

        columns = enum.Enum("Columns", columns)
        for column in columns:
            cell = self.choices_sheet.cell(
                row=1, column=column.value, value=column.name
            )
            cell.font = self.bold_font
            self._add_comment_to_column(column, cell)

        choices_queryset = self.get_choices()
        for choice_group in choices_queryset:
            for choice in choice_group.choices.all().order_by("order", "id"):
                row += 1
                self.choices_sheet.cell(
                    row=row,
                    column=columns["choice_list"].value,
                    value=choice_group.name,
                )
                self.choices_sheet.cell(
                    row=row, column=columns["name"].value, value=choice.name
                )

                if self.languages:
                    if "en" in self.language_values:
                        value, display = ("en", "English")
                        self.choices_sheet.cell(
                            row=row,
                            column=columns[f"label::{display} ({value})"].value,
                            value=choice.label,
                        )
                    for translation in choice.translations.all():
                        value = translation.language
                        display = self.ENGLISH_LANGUAGES.get(value, value)
                        self.choices_sheet.cell(
                            row=row,
                            column=columns[f"label::{display} ({value})"].value,
                            value=translation.label,
                        )

    def add_suffixes(self):
        columns = [
            "name",
            "description",
            "type",
            "choicelist",
            "suffix",
        ]
        columns = enum.Enum("Columns", columns)
        self._fill_column_header(columns, self.suffixes_sheet)
        suffixes_queryset = self.get_suffixes(self.suffixes_ids)
        for i, suffix in enumerate(suffixes_queryset, 2):
            self.suffixes_sheet.cell(
                row=i,
                column=columns["name"].value,
                value=suffix.name,
            )
            self.suffixes_sheet.cell(
                row=i,
                column=columns["description"].value,
                value=suffix.description,
            )
            self.suffixes_sheet.cell(
                row=i,
                column=columns["type"].value,
                value=suffix.type,
            )
            if suffix.choices:
                self.suffixes_sheet.cell(
                    row=i,
                    column=columns["choicelist"].value,
                    value=suffix.choices.name,
                )
            nested_suffixes = ", ".join([s.name for s in suffix.nested_suffixes.all()])
            if nested_suffixes:
                self.suffixes_sheet.cell(
                    row=i,
                    column=columns["suffix"].value,
                    value=nested_suffixes,
                )

    def add_recall_periods(self):
        columns = ["name", "description"]
        columns = enum.Enum("Columns", columns)
        self._fill_column_header(columns, self.recall_period_sheet)
        recall_period_queryset = self.get_recall_periods(self.recall_period_ids)
        for i, recall_period in enumerate(recall_period_queryset, 2):
            self.recall_period_sheet.cell(
                row=i,
                column=columns["name"].value,
                value=recall_period.name,
            )
            self.recall_period_sheet.cell(
                row=i,
                column=columns["description"].value,
                value=recall_period.description,
            )

    def add_required_groups(self):
        columns = ["submodule_name", "suffix1", "suffix2", "recall_period"]
        columns = enum.Enum("Columns", columns)
        self._fill_column_header(columns, self.required_groups_sheet)
        required_groups_queryset = self.get_required_groups(self.required_group_ids)
        for i, group in enumerate(required_groups_queryset, 2):
            self.required_groups_sheet.cell(
                row=i,
                column=columns["submodule_name"].value,
                value=group.submodule.name,
            )
            self.required_groups_sheet.cell(
                row=i,
                column=columns["suffix1"].value,
                value=suffix.name if (suffix := group.required_suffix) else None,
            )
            self.required_groups_sheet.cell(
                row=i,
                column=columns["suffix2"].value,
                value=suffix.name if (suffix := group.required_nested_suffix) else None,
            )
            self.required_groups_sheet.cell(
                row=i,
                column=columns["recall_period"].value,
                value=(
                    recall_period.name
                    if (recall_period := group.required_recall_period)
                    else None
                ),
            )

    def add_indicators(self):
        columns = ["indicator_name", "indicator_label", "question_name"]
        columns = enum.Enum("Columns", columns)
        self._fill_column_header(columns, self.indicators_sheet)
        indicator_queryset = self.get_indicators(self.indicator_ids)
        row = 2
        for indicator in indicator_queryset:
            for question in self.order_indicator_base_questions(
                indicator.questions.filter(id__in=self.base_question_ids)
            ):
                self.indicators_sheet.cell(
                    row=row,
                    column=columns["indicator_name"].value,
                    value=indicator.name,
                )
                self.indicators_sheet.cell(
                    row=row,
                    column=columns["indicator_label"].value,
                    value=indicator.label,
                )
                self.indicators_sheet.cell(
                    row=row,
                    column=columns["question_name"].value,
                    value=question.name,
                )
                row += 1

    def _fill_column_header(self, columns, sheet):
        for column in columns:
            cell = sheet.cell(
                row=1,
                column=column.value,
                value=column.name,
            )
            cell.font = self.bold_font
            self._add_comment_to_column(column, cell)

    def add_questions(self, question_queryset, skip_required_groups=False):
        for question in question_queryset:
            self.add_question_row(question, skip_required_groups)

    def add_from_submodules(self, submodule_queryset):
        root_question_ids = set(
            submodule_queryset.exclude(root_questions=None).values_list(
                "root_questions", flat=True
            )
        )
        root_questions = RootQuestion.objects.filter(id__in=root_question_ids)
        for root_question in root_questions:
            self.add_question_row(root_question, skip_required_groups=True)
            self.add_questions(
                root_question.sub_questions.all(),
                skip_required_groups=True,
            )
        for submodule in submodule_queryset:
            required_group_ids = submodule.required_groups.values_list("id", flat=True)
            self.required_group_ids.update(required_group_ids)

    def fill_sheets(self):
        self.add_choices()
        self.add_suffixes()
        self.add_recall_periods()
        self.add_required_groups()
        self.add_indicators()

    def set_values_from_questions(self, base_question_queryset):
        self._set_question_ids_from_questions(base_question_queryset)
        self._set_repeat_section_ids_from_questions(base_question_queryset)
        self._set_choice_ids_from_questions(base_question_queryset)
        self._set_languages_from_questions()
        self._generate_columns()

    def _set_question_ids_from_questions(self, base_question_queryset):
        self.root_question_ids = (
            base_question_queryset.exclude(root_question=None)
            .values_list("root_question", flat=True)
            .distinct()
        )
        self.sub_question_ids = (
            base_question_queryset.exclude(sub_question=None)
            .values_list("sub_question", flat=True)
            .distinct()
        )

    def _set_choice_ids_from_questions(self, base_question_queryset):
        root_question_choices = list(
            base_question_queryset.exclude(root_question__choices=None).values_list(
                "root_question__choices", flat=True
            )
        )
        suffix_choices = list(
            base_question_queryset.exclude(
                sub_question__suffix__choices=None
            ).values_list("sub_question__suffix__choices", flat=True)
        )
        suffix_2_choices = list(
            base_question_queryset.exclude(
                sub_question__suffix_2__choices=None
            ).values_list("sub_question__suffix_2__choices", flat=True)
        )

        self.choice_groups = list(
            set(root_question_choices + suffix_choices + suffix_2_choices)
        )

        self.choices_ids = (
            Choice.objects.filter(choice_group__in=self.choice_groups)
            .distinct()
            .values_list("id")
        )

    def _set_repeat_section_ids_from_questions(self, base_question_queryset):
        repeat_section_ids = list(
            base_question_queryset.exclude(repeat_section=None).values_list(
                "repeat_section", flat=True
            )
        ) + list(
            base_question_queryset.exclude(repeat_sections=None).values_list(
                "repeat_sections", flat=True
            )
        )
        self.repeat_section_ids.update(repeat_section_ids)

    def _set_languages_from_questions(self):
        root_languages = list(
            RootQuestionTranslation.objects.filter(
                root_question__in=self.root_question_ids
            ).values_list("language", flat=True)
        )
        sub_languages = list(
            SubQuestionTranslation.objects.filter(
                sub_question__in=self.sub_question_ids
            ).values_list("language", flat=True)
        )
        choice_languages = list(
            ChoiceTranslation.objects.filter(choice__in=self.choices_ids).values_list(
                "language", flat=True
            )
        )
        repeat_section_languages = list(
            RepeatSectionTranslation.objects.filter(
                repeat_section__in=self.repeat_section_ids
            ).values_list("language", flat=True)
        )

        language_set = set(
            root_languages + sub_languages + choice_languages + repeat_section_languages
        )
        language_set.discard("en")

        self.language_values = ["en"] + list(language_set)
        self.languages = [
            (language, self.ENGLISH_LANGUAGES.get(language, language))
            for language in self.language_values
        ]

    def generate_from_questions(self, base_question_queryset):
        self.set_values_from_questions(base_question_queryset)
        self.prepare_document()
        self.add_questions(base_question_queryset)
        self.fill_sheets()

        return save_virtual_workbook(self.wb)

    def generate_from_submodules(self, submodule_queryset):
        root_questions_queryset = RootQuestion.objects.filter(
            submodule__in=submodule_queryset
        )
        base_question_queryset = BaseQuestion.objects.filter(
            root_question__in=root_questions_queryset
        ) | BaseQuestion.objects.filter(
            sub_question__root_question__in=root_questions_queryset
        )

        self.set_values_from_questions(
            self.get_optimized_base_question_qs(base_question_queryset)
        )
        self.prepare_document()
        self.add_from_submodules(submodule_queryset)
        self.fill_sheets()
        return save_virtual_workbook(self.wb)

    def generate_from_modules(self, modules_queryset):
        submodule_queryset = Submodule.objects.filter(module__in=modules_queryset)
        root_questions_queryset = RootQuestion.objects.filter(
            submodule__in=submodule_queryset
        )
        base_question_queryset = BaseQuestion.objects.filter(
            root_question__in=root_questions_queryset
        ) | BaseQuestion.objects.filter(
            sub_question__root_question__in=root_questions_queryset
        )

        self.set_values_from_questions(
            self.get_optimized_base_question_qs(base_question_queryset)
        )
        self.prepare_document()
        for module in modules_queryset:
            self.add_from_submodules(module.submodules.all())
        self.fill_sheets()
        return save_virtual_workbook(self.wb)

    def generate_from_indicators(self, indicators_queryset):
        base_question_queryset = self.order_indicator_base_questions(
            BaseQuestion.objects.filter(
                id__in=indicators_queryset.values_list("questions__id", flat=True)
            )
        )
        self.set_values_from_questions(
            self.get_optimized_base_question_qs(base_question_queryset)
        )
        self.prepare_document()
        self.add_questions(self.get_optimized_base_question_qs(base_question_queryset))
        self.fill_sheets()
        return save_virtual_workbook(self.wb)

    def generate_template(self):
        self.IS_TEMPLATE = True
        self._generate_columns()
        self.prepare_document()
        self.fill_sheets()
        return save_virtual_workbook(self.wb)

    @staticmethod
    def get_optimized_submodule_qs(submodule_queryset):
        sub_questions_prefetch = Prefetch(
            "sub_questions",
            queryset=SubQuestion.objects.select_related(
                "suffix__choices", "suffix_2__choices", "recall_period"
            ).prefetch_related("translations", "base_question__repeat_sections"),
        )

        root_questions_prefetch = Prefetch(
            "root_questions",
            queryset=RootQuestion.objects.prefetch_related(
                sub_questions_prefetch, "translations", "base_question__repeat_sections"
            ).select_related("choices"),
        )

        return submodule_queryset.prefetch_related(
            root_questions_prefetch,
        ).distinct()

    @staticmethod
    def get_optimized_base_question_qs(queryset):
        return queryset.prefetch_related(
            "repeat_sections",
            "root_question__translations",
            "sub_question__translations",
        ).select_related(
            "root_question__choices",
            "sub_question__suffix__choices",
            "sub_question__suffix_2__choices",
            "sub_question__recall_period",
        )

    @staticmethod
    def order_indicator_base_questions(base_question_queryset):
        return base_question_queryset.annotate(
            submodule=Coalesce(
                "root_question__submodule",
                "sub_question__root_question__submodule",
            )
        ).order_by(
            "submodule",
            "order",
        )

    def prepare_document(self):
        for index in string.ascii_uppercase:
            self.questions_sheet.column_dimensions[index].width = 20
        self.add_columns()

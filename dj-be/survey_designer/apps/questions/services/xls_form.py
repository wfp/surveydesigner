import enum
import os
import secrets
import string

from django.db.models import Prefetch, Q
from modules.models import (
    Indicator,
    Module,
    ModuleTranslation,
    Submodule,
    SubmoduleTranslation,
)
from openpyxl import Workbook
from openpyxl.styles import Font
from questions.const import QuestionType
from questions.models import (
    BaseQuestion,
    Choice,
    ChoiceGroup,
    ChoiceTranslation,
    RepeatSection,
    RepeatSectionTranslation,
    RootQuestion,
    RootQuestionTranslation,
    SubQuestion,
    SubQuestionTranslation,
)

from .workbook import save_virtual_workbook


class XLSForm:
    STATIC_COLUMNS = (
        "type",
        "name",
        "label",
        "hint",
        "repeat_count",
        "constraint",
        "constraint_message",
        "required",  # TODO: This is already here?
        "required_message",  # TODO: What is this?
        "relevant",
        "choice_filter",
        "calculation",
        "appearance",
        "parameters",
        "default",
        "disabled",
        "read_only",
    )
    LANGUAGE_COLUMNS = ("label", "hint", "constraint_message", "required_message")
    METADATA = ("start", "end", "today", "deviceid")
    ENGLISH_LANGUAGES = {
        "en": "English",
        "fr": "French",
        "es": "Spanish",
        "ar": "Arabic",
        "zh-cn": "Chinese",
        "ru": "Russian",
        "pt": "Portuguese",
    }

    def __init__(
        self,
        name,
        submodule_ids,
        sub_question_ids,
        submodules_order,
        languages=None,
        indicators=None,
        protected=False,
    ):
        self.name = name
        self.protected = protected
        self.id_name = secrets.token_urlsafe()
        self.language_values = languages or ()
        self.languages = [
            (language, self.ENGLISH_LANGUAGES.get(language, language))
            for language in self.language_values
        ]
        self.indicators = self.get_indicators(indicators or [])
        self.submodules_order = submodules_order
        self.selected_submodule_ids = submodule_ids
        self.submodule_ids = self.get_submodule_ids()
        self.module_ids = self.get_module_ids()
        self.modules = self.get_modules()
        self.current_module = None
        self.sub_question_ids = sub_question_ids
        self.columns = self.generate_columns()
        self.current_row_index = 1
        self.choices_ids = set()
        self.bold_font = Font(bold=True)
        self.processed_questions = set()
        self.external_files = {}

        self.wb = Workbook()
        self.survey_sheet = self.wb.active
        self.survey_sheet.title = "survey"
        self.choices_sheet = self.wb.create_sheet(title="choices")
        self.settings_sheet = self.wb.create_sheet(title="settings")

    @staticmethod
    def get_indicators(ids):
        return Indicator.objects.filter(id__in=ids).prefetch_related(
            "questions__root_question__submodule",
            "questions__sub_question__root_question__submodule",
        )

    def get_submodule_ids(self):
        indicator_submodule_ids = list(
            self.indicators.values_list(
                "questions__root_question__submodule", flat=True
            )
        )
        return [
            s_id
            for s_id in self.submodules_order
            if s_id in indicator_submodule_ids + self.selected_submodule_ids
        ]

    def get_module_ids(self):
        module_ids = [
            i[1]
            for i in sorted(
                Submodule.objects.filter(id__in=self.submodule_ids).values_list(
                    "id", "module__id"
                ),
                key=lambda x: self.submodule_ids.index(x[0]),
            )
        ]
        return sorted(set(module_ids), key=lambda m_id: module_ids.index(m_id))

    def get_modules(self):
        sub_questions_translations_prefetch = Prefetch(
            "translations",
            queryset=SubQuestionTranslation.objects.filter(
                language__in=self.language_values
            ),
        )

        sub_questions_prefetch = Prefetch(
            "sub_questions",
            queryset=SubQuestion.objects.select_related(
                "suffix__choices", "suffix_2__choices", "recall_period"
            ).prefetch_related(sub_questions_translations_prefetch),
        )

        root_questions_translations_prefetch = Prefetch(
            "translations",
            queryset=RootQuestionTranslation.objects.filter(
                language__in=self.language_values
            ),
        )

        root_questions_prefetch = Prefetch(
            "root_questions",
            queryset=RootQuestion.objects.prefetch_related(
                sub_questions_prefetch, root_questions_translations_prefetch
            )
            .select_related("choices")
            .filter(base_question__repeat_sections=None),
        )

        repeat_sections_translations_prefetch = Prefetch(
            "translations",
            queryset=RepeatSectionTranslation.objects.filter(
                language__in=self.language_values
            ),
        )

        repeat_sections_prefetch = Prefetch(
            "repeat_sections",
            queryset=RepeatSection.objects.prefetch_related(
                repeat_sections_translations_prefetch,
                "questions",
                "repeat_count_dependencies",
            ),
        )

        submodule_translations_prefetch = Prefetch(
            "translations",
            queryset=SubmoduleTranslation.objects.filter(
                language__in=self.language_values
            ),
        )

        submodules_prefetch = Prefetch(
            "submodules",
            queryset=Submodule.objects.filter(id__in=self.submodule_ids)
            .prefetch_related(
                root_questions_prefetch,
                submodule_translations_prefetch,
                repeat_sections_prefetch,
            )
            .distinct(),
        )

        module_translations_prefetch = Prefetch(
            "translations",
            queryset=ModuleTranslation.objects.filter(
                language__in=self.language_values
            ),
        )

        modules = Module.objects.filter(id__in=self.module_ids).prefetch_related(
            module_translations_prefetch,
            submodules_prefetch,
        )

        return sorted(modules, key=lambda module: self.module_ids.index(module.id))

    def get_choices(self, ids):
        choices_translations_prefetch = Prefetch(
            "translations",
            queryset=ChoiceTranslation.objects.filter(
                language__in=self.language_values
            ),
        )
        choices_prefetch = Prefetch(
            "choices",
            queryset=Choice.objects.filter(is_active=True)
            .prefetch_related(choices_translations_prefetch)
            .order_by("order", "id"),
        )
        return ChoiceGroup.objects.filter(id__in=ids).prefetch_related(choices_prefetch)

    @staticmethod
    def get_language_column(name, language):
        value, display = language
        return f"{name}::{display} ({value})"

    def get_column_value(self, name, language=None):
        key = self.get_language_column(name, language) if language else name
        try:
            return self.columns[key].value
        except KeyError:
            # TODO: add logging
            return

    def generate_columns(self):
        columns = []

        for column in self.STATIC_COLUMNS:
            if self.languages and column in self.LANGUAGE_COLUMNS:
                for language in self.languages:
                    columns.append(self.get_language_column(column, language))
            else:
                columns.append(column)

        return enum.Enum("Columns", columns)

    def increment_row_index(self):
        self.current_row_index += 1

    def fill_cell(self, column_name, value, bold=False, language: tuple = None):
        column = self.get_column_value(column_name, language=language)

        if not column:
            # TODO: add logging
            return

        cell = self.survey_sheet.cell(
            row=self.current_row_index, column=column, value=value
        )

        if bold:
            cell.font = self.bold_font
        return cell

    def add_metadata(self):
        for meta in self.METADATA:
            self.increment_row_index()
            self.fill_cell("type", meta)
            self.fill_cell("name", meta)

    def _track_external_file(self, question):
        if question.type not in (
            QuestionType.SELECT_ONE_FROM_FILE,
            QuestionType.SELECT_MULTIPLE_FROM_FILE,
        ):
            return

        choices_file = None
        if hasattr(question, "choices_file") and question.choices_file:
            choices_file = question.choices_file
        else:
            choices = getattr(question, "choices", None)
            if choices is not None and hasattr(choices, "csv_file"):
                choices_file = choices

        if not choices_file or not getattr(choices_file, "csv_file", None):
            return

        file_field = choices_file.csv_file
        file_name = os.path.basename(file_field.name)
        if file_name and file_name not in self.external_files:
            self.external_files[file_name] = file_field

    def add_question_row(self, question):
        self.increment_row_index()
        self.fill_cell("type", question.get_xls_type())
        self.fill_cell("name", question.name)
        self.fill_cell("constraint", question.constraint)
        self.fill_cell("relevant", question.relevant)
        self.fill_cell("choice_filter", question.choice_filter)
        self.fill_cell("appearance", question.appearance)
        self.fill_cell("default", question.default)
        self.fill_cell("read_only", question.read_only)
        self.fill_cell("required", question.required)
        self.fill_cell("disabled", question.disabled)
        self.fill_cell("calculation", question.calculation)
        self.fill_cell("parameters", question.parameters)

        if self.languages:
            if "en" in self.language_values:
                language = ("en", "English")
                self.fill_cell("label", question.label, language=language)
                self.fill_cell("hint", question.hint, language=language)
                self.fill_cell(
                    "constraint_message", question.constraint_message, language=language
                )

            for translation in question.translations.all():
                value = translation.language
                language = (value, self.ENGLISH_LANGUAGES.get(value, value))
                self.fill_cell("label", translation.label, language=language)
                self.fill_cell("hint", translation.hint, language=language)

            for translation in question.constraint_translations.all():
                value = translation.language
                language = (value, self.ENGLISH_LANGUAGES.get(value, value))
                self.fill_cell(
                    "constraint_message", translation.label, language=language
                )
        else:
            self.fill_cell("label", question.label)
            self.fill_cell("hint", question.hint)
            self.fill_cell("constraint_message", question.constraint_message)

        if question.choices:
            self.choices_ids.add(question.choices.id)

    def add_repeat_section(self, repeat_section):
        self.increment_row_index()
        self.fill_cell("type", "begin_repeat", bold=True)
        self.fill_cell("name", repeat_section.name)
        self.fill_cell("relevant", repeat_section.relevant)
        self.fill_cell("repeat_count", repeat_section.repeat_count)

        if self.languages:
            if "en" in self.language_values:
                language = ("en", "English")
                self.fill_cell("label", repeat_section.label, language=language)
            for translation in repeat_section.translations.all():
                value = translation.language
                language = (value, self.ENGLISH_LANGUAGES.get(value, value))
                self.fill_cell("label", translation.label, language=language)
        else:
            self.fill_cell("label", repeat_section.label)

        selected_questions_query = Q(
            Q(sub_question_id__isnull=True)
            | Q(sub_question_id__in=self.sub_question_ids)
        )
        selected_suffix_ids = set(
            repeat_section.questions.filter(
                selected_questions_query,
            )
            .exclude(sub_question__suffix__name="_oth")
            .values_list("sub_question__suffix_id", flat=True)
        )
        for base_question in repeat_section.questions.filter(
            selected_questions_query
            | Q(sub_question__suffix__name="_oth")
            | (
                Q(sub_question__suffix_id__in=selected_suffix_ids)
                & Q(sub_question__suffix_2__name="_oth")
            )
        ).order_by("order"):
            self.add_question_row(base_question.instance)

        self.increment_row_index()
        self.fill_cell("type", "end_repeat", bold=True)

    def add_settings(self):
        columns = ["form_title", "form_id"]
        columns = enum.Enum("Columns", columns)
        for column in columns:
            self.settings_sheet.cell(row=1, column=column.value, value=column.name)

        self.settings_sheet.cell(
            row=2, column=columns["form_title"].value, value=self.name
        )
        self.settings_sheet.cell(
            row=2, column=columns["form_id"].value, value=self.id_name
        )

    def add_choices(self):
        columns = ["list_name", "name", "choice_filter_name"]
        if self.languages:
            for language in self.languages:
                value, display = language
                columns.append(f"label::{display} ({value})")
        else:
            columns.append("label")

        row = 1

        columns = enum.Enum("Columns", columns)

        for column in columns:
            self.choices_sheet.cell(row=1, column=column.value, value=column.name)

        unordered_choices_queryset = self.get_choices(self.choices_ids)
        objects = unordered_choices_queryset.in_bulk(self.choices_ids)
        choices_queryset = [objects[id] for id in self.choices_ids if id in objects]
        for choice_group in choices_queryset:
            for choice in choice_group.choices.filter(is_active=True):
                row += 1
                self.choices_sheet.cell(
                    row=row, column=columns["list_name"].value, value=choice_group.name
                )
                self.choices_sheet.cell(
                    row=row, column=columns["name"].value, value=choice.name
                )
                self.choices_sheet.cell(
                    row=row,
                    column=columns["choice_filter_name"].value,
                    value=(
                        choice.choice_filter_name.name
                        if choice.choice_filter_name
                        else ""
                    ),
                )
                if self.languages:
                    if "en" in self.language_values:
                        value, display = ("en", "English")
                        self.choices_sheet.cell(
                            row=row,
                            column=columns[f"label::{display} ({value})"].value,
                            value=choice.label,
                        )
                    for translation in choice.translations.all():
                        value = translation.language
                        display = self.ENGLISH_LANGUAGES.get(value, value)
                        self.choices_sheet.cell(
                            row=row,
                            column=columns[f"label::{display} ({value})"].value,
                            value=translation.label,
                        )
                else:
                    self.choices_sheet.cell(
                        row=row, column=columns["label"].value, value=choice.label
                    )

    def add_columns(self):
        for column in self.columns:
            cell = self.survey_sheet.cell(row=1, column=column.value, value=column.name)
            cell.font = self.bold_font

    def fill_submodule_start(self, submodule):
        self.fill_cell("type", "begin_group", bold=True)
        self.fill_cell("name", f"{submodule.name}_submodule")
        self.fill_cell("appearance", submodule.appearance)
        self.fill_cell("relevant", submodule.relevant)

        if self.languages:
            if "en" in self.language_values:
                language = ("en", "English")
                self.fill_cell("label", submodule.label, language=language)

            for translation in submodule.translations.all():
                value = translation.language
                language = (value, self.ENGLISH_LANGUAGES.get(value, value))
                self.fill_cell("label", translation.label, language=language)
        else:
            self.fill_cell("label", submodule.label)

    def fill_module_start(self, module):
        self.fill_cell("type", "begin_group", bold=True)
        self.fill_cell("name", f"{module.name}_module")
        self.fill_cell("relevant", module.relevant)
        if self.languages:
            if "en" in self.language_values:
                language = ("en", "English")
                self.fill_cell("label", module.label, language=language)

            for translation in module.translations.all():
                value = translation.language
                language = (value, self.ENGLISH_LANGUAGES.get(value, value))
                self.fill_cell("label", translation.label, language=language)
        else:
            self.fill_cell("label", module.label)

    def add_questions(self):
        for module in self.modules:
            self.increment_row_index()
            self.fill_module_start(module)

            submodules = sorted(
                module.submodules.all(),
                key=lambda submodule: self.submodule_ids.index(submodule.id),
            )
            for submodule in submodules:
                root_questions = RootQuestion.objects.filter(
                    submodule=submodule
                ).exists()
                if not root_questions:
                    continue
                is_selected = submodule.id in self.selected_submodule_ids
                if is_selected:
                    self.add_selected_submodule(submodule)
                if not is_selected:
                    # Submodule was pulled in via indicator selection
                    indicator_submodule_ids = self.indicators.values_list(
                        "questions__root_question__submodule", flat=True
                    ).distinct()
                    if submodule.id in indicator_submodule_ids:
                        self.add_indicator_submodule(submodule)

            self.increment_row_index()
            self.fill_cell("type", "end_group", bold=True)

    def add_selected_submodule(self, submodule):
        self.increment_row_index()
        self.fill_submodule_start(submodule)

        repeats = BaseQuestion.objects.filter(repeat_section__submodule=submodule)
        roots = BaseQuestion.objects.filter(
            root_question__submodule=submodule,
        ).exclude(
            id__in=repeats.values_list("repeat_section__questions", flat=True),
        )

        base_questions = (roots | repeats).order_by("order")

        for question in base_questions:
            question_instance = question.instance
            if isinstance(question_instance, RepeatSection):
                self.add_repeat_section(question_instance)
            else:
                self.add_root_question(question_instance)

        self.increment_row_index()
        self.fill_cell("type", "end_group", bold=True)

    def add_indicator_submodule(self, submodule):
        self.increment_row_index()
        self.fill_submodule_start(submodule)

        indicator_repeat_section_ids_for_submodule = self.indicators.filter(
            questions__repeat_section__submodule=submodule,
        ).values_list("questions__repeat_section", flat=True)
        repeats = BaseQuestion.objects.filter(
            repeat_section__id__in=indicator_repeat_section_ids_for_submodule,
        )

        indicator_root_question_ids_for_submodule = self.indicators.filter(
            questions__root_question__submodule=submodule,
        ).values_list("questions__root_question", flat=True)
        roots = BaseQuestion.objects.filter(
            root_question__id__in=indicator_root_question_ids_for_submodule,
        ).exclude(
            id__in=repeats.values_list("repeat_section__questions", flat=True),
        )
        base_questions = (roots | repeats).order_by("order")
        for question in base_questions:
            question_instance = question.instance
            if isinstance(question_instance, RepeatSection):
                self.add_repeat_section(question_instance)
            else:
                self.add_root_question(question_instance)

        self.increment_row_index()
        self.fill_cell("type", "end_group", bold=True)

    def add_root_question(self, root_question):
        self.processed_questions.add(root_question.name)
        self.add_question_row(root_question)
        self._track_external_file(root_question)

        selected_suffixes = []
        for sub_question in root_question.sub_questions.order_by(
            "base_question__order",
        ):
            suffix = sub_question.suffix
            suffix_2 = sub_question.suffix_2
            is_selected = sub_question.id in self.sub_question_ids

            if (
                is_selected
                or (suffix and suffix.name == "_oth")
                or (
                    suffix in selected_suffixes and suffix_2 and suffix_2.name == "_oth"
                )
            ):
                self.processed_questions.add(sub_question.name)
                self.add_question_row(sub_question)
                self._track_external_file(sub_question)

            if is_selected and suffix and suffix.name != "_oth":
                selected_suffixes.append(sub_question.suffix)

    def generate(self):
        for index in string.ascii_uppercase:
            self.survey_sheet.column_dimensions[index].width = 20
        self.add_columns()
        self.add_questions()
        self.add_metadata()
        self.add_choices()
        self.add_settings()
        if self.protected:
            self.protect_workbook()
        return save_virtual_workbook(self.wb)

    def protect_workbook(self):
        for sheet in self.wb.worksheets:
            sheet.protection.password = secrets.token_hex(16)
            sheet.protection.enable()
